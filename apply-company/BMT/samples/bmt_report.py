"""
BMT Excel report writer.

- Summary
- Details
- SN Validation
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bmt_excel_utils import parse_numeric_qty


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
STATUS_FILLS = {
    "CRITICAL": PatternFill(fill_type="solid", fgColor="FFCCCC"),
    "WARN": PatternFill(fill_type="solid", fgColor="FFE0B2"),
    "REVIEW": PatternFill(fill_type="solid", fgColor="FFF9C4"),
    "OK": PatternFill(fill_type="solid", fgColor="C8E6C9"),
    "SKIP": PatternFill(fill_type="solid", fgColor="E0E0E0"),
}
STATUS_TO_LEVEL = {
    "exact": "OK",
    "ok": "OK",
    "qty_mismatch": "WARN",
    "fuzzy": "WARN",
    "prefix": "WARN",
    "synonym": "REVIEW",
    "balloon_image_only": "WARN",
    "balloon_table_only_critical": "CRITICAL",
    "drawing_only": "CRITICAL",
    "bom_only": "CRITICAL",
    "skipped": "SKIP",
}


def _apply_header_style(row) -> None:
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(48, len(value) + 2))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _apply_status_fill(cell, value: object) -> None:
    text = str(value or "").strip()
    level = STATUS_TO_LEVEL.get(text.lower(), text.upper())
    if level in STATUS_FILLS:
        cell.fill = STATUS_FILLS[level]


def _qty_delta_text(drawing_qty: object, erp_qty: object) -> object:
    drawing_text = str(drawing_qty or "").strip()
    erp_text = str(erp_qty or "").strip()
    if not drawing_text and not erp_text:
        return ""
    return parse_numeric_qty(drawing_qty) - parse_numeric_qty(erp_qty)


def _build_mismatch_rows(detail_rows: Iterable[Dict[str, object]]) -> List[Dict[str, object]]:
    mismatch_rows: List[Dict[str, object]] = []
    for detail in detail_rows:
        issue_type = str(detail.get("verdict", ""))
        if issue_type == "exact":
            continue
        severity = str(detail.get("severity", "")).strip().upper() or STATUS_TO_LEVEL.get(issue_type, "")
        drawing_item = (
            detail.get("drawing_part_no")
            or detail.get("part_list_code")
            or detail.get("part_list_drawing_no")
            or ""
        )
        mismatch_rows.append(
            {
                "severity": severity,
                "issue_type": issue_type,
                "dataset": detail.get("dataset", ""),
                "section": detail.get("section", ""),
                "page": detail.get("page", ""),
                "drawing_item": drawing_item,
                "erp_item": detail.get("erp_child_item", ""),
                "drawing_qty": detail.get("drawing_qty", ""),
                "erp_qty": detail.get("erp_qty", ""),
                "delta": _qty_delta_text(detail.get("drawing_qty", ""), detail.get("erp_qty", "")),
                "reason": detail.get("detail", ""),
                "review_status": detail.get("review_status", "미확인"),
                "reviewer": detail.get("reviewer", ""),
                "comment": detail.get("comment", ""),
            }
        )
    return mismatch_rows


def _number_list_text(values: object) -> str:
    if not values:
        return "-"
    if isinstance(values, str):
        return values
    return ", ".join(str(value) for value in values)


def _write_summary(ws, summary_rows: List[Dict[str, object]]) -> None:
    ws["A1"] = "BMT Drawing-BOM Verification Report"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = "Generated"
    ws["B2"] = datetime.now().isoformat(timespec="seconds")

    headers = [
        "Dataset",
        "Records",
        "Exact",
        "Qty Mismatch",
        "Prefix",
        "Fuzzy",
        "Synonym",
        "Drawing Only",
        "BOM Only",
        "Status",
        "Notes",
    ]
    header_row = 4
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=column_index, value=header)
    _apply_header_style(ws[header_row])

    for row_index, summary in enumerate(summary_rows, start=header_row + 1):
        ws.cell(row=row_index, column=1, value=summary.get("dataset", ""))
        ws.cell(row=row_index, column=2, value=summary.get("records", 0))
        ws.cell(row=row_index, column=3, value=summary.get("exact", 0))
        ws.cell(row=row_index, column=4, value=summary.get("qty_mismatch", 0))
        ws.cell(row=row_index, column=5, value=summary.get("prefix", 0))
        ws.cell(row=row_index, column=6, value=summary.get("fuzzy", 0))
        ws.cell(row=row_index, column=7, value=summary.get("synonym", 0))
        ws.cell(row=row_index, column=8, value=summary.get("drawing_only", 0))
        ws.cell(row=row_index, column=9, value=summary.get("bom_only", 0))
        ws.cell(row=row_index, column=10, value=summary.get("status", ""))
        ws.cell(row=row_index, column=11, value=summary.get("notes", ""))
        _apply_status_fill(ws.cell(row=row_index, column=10), summary.get("status", ""))

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{max(4, header_row + len(summary_rows))}"
    _autosize(ws)


def _write_details(ws, detail_rows: Iterable[Dict[str, object]]) -> None:
    headers = [
        "Dataset",
        "Section",
        "Page",
        "Drawing Part No.",
        "Part List Code",
        "Part List Drawing No.",
        "ERP Child Item",
        "Drawing Qty",
        "ERP Qty",
        "Qty Compare",
        "Verdict",
        "Detail",
        "Review Status",
        "Reviewer",
        "Comment",
    ]
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column_index, value=header)
    _apply_header_style(ws[1])

    for row_index, detail in enumerate(detail_rows, start=2):
        ws.cell(row=row_index, column=1, value=detail.get("dataset", ""))
        ws.cell(row=row_index, column=2, value=detail.get("section", ""))
        ws.cell(row=row_index, column=3, value=detail.get("page", ""))
        ws.cell(row=row_index, column=4, value=detail.get("drawing_part_no", ""))
        ws.cell(row=row_index, column=5, value=detail.get("part_list_code", ""))
        ws.cell(row=row_index, column=6, value=detail.get("part_list_drawing_no", ""))
        ws.cell(row=row_index, column=7, value=detail.get("erp_child_item", ""))
        ws.cell(row=row_index, column=8, value=detail.get("drawing_qty", ""))
        ws.cell(row=row_index, column=9, value=detail.get("erp_qty", ""))
        ws.cell(row=row_index, column=10, value=detail.get("qty_compare", ""))
        ws.cell(row=row_index, column=11, value=detail.get("verdict", ""))
        ws.cell(row=row_index, column=12, value=detail.get("detail", ""))
        ws.cell(row=row_index, column=13, value=detail.get("review_status", "미확인"))
        ws.cell(row=row_index, column=14, value=detail.get("reviewer", ""))
        ws.cell(row=row_index, column=15, value=detail.get("comment", ""))
        _apply_status_fill(ws.cell(row=row_index, column=11), detail.get("verdict", ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{max(1, ws.max_row)}"
    _autosize(ws)


def _write_mismatches(ws, mismatch_rows: List[Dict[str, object]]) -> None:
    headers = [
        "Severity",
        "Issue Type",
        "Dataset",
        "Section",
        "Page",
        "Drawing Item",
        "ERP Item",
        "Drawing Qty",
        "ERP Qty",
        "Delta",
        "Reason",
        "Review Status",
        "Reviewer",
        "Comment",
    ]
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column_index, value=header)
    _apply_header_style(ws[1])

    for row_index, mismatch in enumerate(mismatch_rows, start=2):
        ws.cell(row=row_index, column=1, value=mismatch.get("severity", ""))
        ws.cell(row=row_index, column=2, value=mismatch.get("issue_type", ""))
        ws.cell(row=row_index, column=3, value=mismatch.get("dataset", ""))
        ws.cell(row=row_index, column=4, value=mismatch.get("section", ""))
        ws.cell(row=row_index, column=5, value=mismatch.get("page", ""))
        ws.cell(row=row_index, column=6, value=mismatch.get("drawing_item", ""))
        ws.cell(row=row_index, column=7, value=mismatch.get("erp_item", ""))
        ws.cell(row=row_index, column=8, value=mismatch.get("drawing_qty", ""))
        ws.cell(row=row_index, column=9, value=mismatch.get("erp_qty", ""))
        ws.cell(row=row_index, column=10, value=mismatch.get("delta", ""))
        ws.cell(row=row_index, column=11, value=mismatch.get("reason", ""))
        ws.cell(row=row_index, column=12, value=mismatch.get("review_status", "미확인"))
        ws.cell(row=row_index, column=13, value=mismatch.get("reviewer", ""))
        ws.cell(row=row_index, column=14, value=mismatch.get("comment", ""))
        _apply_status_fill(ws.cell(row=row_index, column=1), mismatch.get("severity", ""))
        _apply_status_fill(ws.cell(row=row_index, column=2), mismatch.get("issue_type", ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max(1, ws.max_row)}"
    _autosize(ws)


def _write_sn_validation(ws, sn_validation: Dict[str, object]) -> None:
    ws["A1"] = "Section Summary"
    ws["A1"].font = Font(size=13, bold=True)
    ws["A2"] = "Expected SN"
    ws["B2"] = sn_validation.get("expected_sn", "")
    ws["D2"] = "PDF"
    ws["E2"] = sn_validation.get("pdf_path", "")

    section_headers = [
        "Section",
        "Pages",
        "Detected SNs",
        "Representative SN",
        "Status",
        "Doc Numbers",
    ]
    section_header_row = 4
    for column_index, header in enumerate(section_headers, start=1):
        ws.cell(row=section_header_row, column=column_index, value=header)
    _apply_header_style(ws[section_header_row])

    row_index = section_header_row + 1
    for section in sn_validation.get("sections", []):
        ws.cell(row=row_index, column=1, value=section.get("section", ""))
        ws.cell(row=row_index, column=2, value=section.get("page_range", ""))
        ws.cell(row=row_index, column=3, value=", ".join(section.get("detected_sns", [])))
        ws.cell(row=row_index, column=4, value=section.get("representative_sn", ""))
        ws.cell(row=row_index, column=5, value=section.get("status", ""))
        ws.cell(row=row_index, column=6, value="\n".join(section.get("doc_numbers", [])))
        row_index += 1

    row_index += 2
    ws.cell(row=row_index, column=1, value="Page Detail")
    ws.cell(row=row_index, column=1).font = Font(size=13, bold=True)
    row_index += 1

    page_headers = ["Page", "Section", "Doc No.", "Detected SN", "Status"]
    for column_index, header in enumerate(page_headers, start=1):
        ws.cell(row=row_index, column=column_index, value=header)
    _apply_header_style(ws[row_index])
    row_index += 1

    for page in sn_validation.get("pages", []):
        ws.cell(row=row_index, column=1, value=page.get("page", ""))
        ws.cell(row=row_index, column=2, value=page.get("section", ""))
        ws.cell(row=row_index, column=3, value=page.get("doc_no", ""))
        ws.cell(row=row_index, column=4, value=page.get("sn", ""))
        ws.cell(row=row_index, column=5, value=page.get("status", ""))
        row_index += 1

    ws.freeze_panes = "A5"
    _autosize(ws)


def _write_image_cross_checks(ws, image_cross_checks: Iterable[Dict[str, object]]) -> None:
    headers = [
        "Page",
        "Balloon Numbers (image)",
        "Table Item Numbers",
        "Image Only",
        "Table Only",
        "Match Count",
        "Status",
    ]
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column_index, value=header)
    _apply_header_style(ws[1])

    for row_index, row in enumerate(image_cross_checks, start=2):
        ws.cell(row=row_index, column=1, value=row.get("page_label") or row.get("page", ""))
        ws.cell(row=row_index, column=2, value=_number_list_text(row.get("balloon_numbers", [])))
        ws.cell(row=row_index, column=3, value=_number_list_text(row.get("table_item_numbers", [])))
        ws.cell(row=row_index, column=4, value=_number_list_text(row.get("image_only", [])))
        ws.cell(row=row_index, column=5, value=_number_list_text(row.get("table_only", [])))
        ws.cell(row=row_index, column=6, value=row.get("match_count", 0))
        ws.cell(row=row_index, column=7, value=row.get("status", ""))
        _apply_status_fill(ws.cell(row=row_index, column=7), row.get("status", ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(1, ws.max_row)}"
    _autosize(ws)


def write_comparison_report(
    report_path: str,
    summary_rows: List[Dict[str, object]],
    detail_rows: List[Dict[str, object]],
    sn_validation: Dict[str, object],
    mismatch_rows: List[Dict[str, object]] | None = None,
    image_cross_checks: List[Dict[str, object]] | None = None,
) -> str:
    """비교 결과를 XLSX로 저장"""
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    detail_ws = workbook.create_sheet("Details")
    mismatch_ws = workbook.create_sheet("Mismatches")
    sn_ws = workbook.create_sheet("SN Validation")
    image_cross_check_ws = workbook.create_sheet("Image Cross-Check")

    _write_summary(summary_ws, summary_rows)
    _write_details(detail_ws, detail_rows)
    _write_mismatches(mismatch_ws, mismatch_rows if mismatch_rows is not None else _build_mismatch_rows(detail_rows))
    _write_sn_validation(sn_ws, sn_validation)
    _write_image_cross_checks(image_cross_check_ws, image_cross_checks or [])

    workbook.save(report_path)
    return report_path

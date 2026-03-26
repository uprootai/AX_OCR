"""
BMT End-to-End 파이프라인
입력: GAR 배치도 이미지 1장 + Part List 엑셀 + ERP BOM 엑셀
출력: BOM 누락 리포트 (Excel)
"""
import numpy as np
from PIL import Image
import cv2
import requests
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# === Config ===
OCR_URLS = {
    'paddle': 'http://localhost:5006/api/v1/ocr',
    'tesseract': 'http://localhost:5008/api/v1/ocr',
}
TAG_PATTERNS = [re.compile(r'^V\d+(-\d+)?$'), re.compile(r'^PT\d+$'), re.compile(r'^TT\d+$'),
    re.compile(r'^FT\d+$'), re.compile(r'^PI\d+$'), re.compile(r'^B\d+$')]
NOISE_PATTERNS = [re.compile(r'^\d'), re.compile(r'GVU'), re.compile(r'WORKSPACE'),
    re.compile(r'ENCLOSURE'), re.compile(r'MED1A'), re.compile(r'DN\d'),
    re.compile(r'^[A-Z]\d$'), re.compile(r'SPEC'), re.compile(r'PRESSURE'),
    re.compile(r','), re.compile(r'FUEL'), re.compile(r'NATURAL'),
    re.compile(r'NITROGEN'), re.compile(r'STPG'), re.compile(r'SUS'), re.compile(r'^R\d')]
CORRECTIONS = {'Z': '7', 'O': '0', 'I': '1', 'S': '5'}


def correct_tag(tag):
    m = re.match(r'^([A-Z]+)(.*)', tag)
    if not m: return tag
    return m.group(1) + ''.join(CORRECTIONS.get(c, c) for c in m.group(2))


def is_valid_tag(t):
    t = t.strip().upper()
    for p in NOISE_PATTERNS:
        if p.search(t): return False
    for p in TAG_PATTERNS:
        if p.match(t): return True
    return False


# === Step 1: 뷰 라벨 검출 ===
def detect_view_labels(image_path):
    with open(image_path, 'rb') as f:
        r = requests.post(OCR_URLS['paddle'], files={'file': ('img.png', f, 'image/png')}, timeout=120).json()
    labels = []
    for det in r.get('detections', []):
        if re.search(r'VIEW', det['text'], re.IGNORECASE):
            bbox = det['bbox']
            labels.append({'text': det['text'].strip(),
                'cx': (bbox[0][0]+bbox[2][0])/2, 'cy': (bbox[0][1]+bbox[2][1])/2})
    return labels


# === Step 2: Min-Content 크롭 ===
def min_content_split(image_path):
    img = Image.open(image_path)
    gray = np.array(img.convert('L'))
    h, w = gray.shape
    _, binary = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)

    labels = detect_view_labels(image_path)
    if not labels: return [{'name': 'FULL', 'bbox': (0, 0, w, h)}]
    labels.sort(key=lambda v: v['cy'])

    # 행 클러스터
    rows = []; cur = [labels[0]]
    for v in labels[1:]:
        if v['cy'] - cur[-1]['cy'] < 100: cur.append(v)
        else: rows.append(cur); cur = [v]
    rows.append(cur)
    for row in rows: row.sort(key=lambda v: v['cx'])

    def find_min_col(y1, y2, xl, xr, win=20):
        band = binary[y1:y2, xl:xr]
        cs = np.sum(band > 0, axis=0)
        if len(cs) < win: return (xl+xr)//2
        s = np.convolve(cs, np.ones(win)/win, mode='valid')
        return xl + np.argmin(s) + win//2

    def find_min_row(x1, x2, yt, yb, win=20):
        band = binary[yt:yb, x1:x2]
        rs = np.sum(band > 0, axis=1)
        if len(rs) < win: return (yt+yb)//2
        s = np.convolve(rs, np.ones(win)/win, mode='valid')
        return yt + np.argmin(s) + win//2

    row_bounds = [find_min_row(0, w, int(np.mean([v['cy'] for v in rows[i]])),
                    int(np.mean([v['cy'] for v in rows[i+1]])))
                  for i in range(len(rows)-1)]

    last_ly = max(v['cy'] for v in rows[-1])
    rm = np.mean(gray, axis=1)
    row_bottom = h; cnt = 0
    for y in range(int(last_ly+150), h):
        if rm[y] > 248: cnt += 1
        else: cnt = 0
        if cnt >= 20: row_bottom = y - 20; break

    crops = []
    for ri, row in enumerate(rows):
        yt = row_bounds[ri-1] if ri > 0 else 0
        yb = row_bounds[ri] if ri < len(row_bounds) else row_bottom
        cb = [find_min_col(yt, yb, int(row[vi]['cx']), int(row[vi+1]['cx']))
              for vi in range(len(row)-1)]
        xe = [0] + cb + [w]
        for vi, v in enumerate(row):
            crops.append({'name': v['text'], 'bbox': (xe[vi], yt, xe[vi+1], yb)})

    if row_bottom < h - 100:
        ly = row_bottom - 30
        crops.append({'name': '3D+LOWER', 'bbox': (0, ly, w, h)})
        mid = w // 2
        crops.append({'name': '3D LEFT', 'bbox': (0, ly, mid, h)})
        crops.append({'name': '3D RIGHT', 'bbox': (mid-100, ly, w, h)})

    return crops


# === Step 3: OCR + TAG 추출 ===
def extract_tags_from_crops(image_path, crops):
    img = Image.open(image_path)
    all_tags = set()
    per_crop = {}

    for crop in crops:
        x1, y1, x2, y2 = crop['bbox']
        roi = img.crop((max(0,x1), max(0,y1), min(img.size[0],x2), min(img.size[1],y2)))
        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
        roi.save(tmp.name); tmp.close()

        crop_tags = set()
        for eng, url in OCR_URLS.items():
            try:
                with open(tmp.name, 'rb') as f:
                    resp = requests.post(url, files={'file': ('c.png', f, 'image/png')}, timeout=60).json()
                texts = [d['text'] for d in resp.get('detections' if eng == 'paddle' else 'texts', [])]
                if eng != 'paddle': texts += [resp.get('full_text', '')]
                for t in texts:
                    tw = t.strip().upper().replace(' ', '')
                    tw = re.sub(r'[|=(){}\[\]]', '', tw)
                    c = correct_tag(tw)
                    if is_valid_tag(c): crop_tags.add(c)
                    if is_valid_tag(tw): crop_tags.add(tw)
            except: pass

        os.unlink(tmp.name)
        per_crop[crop['name']] = sorted(crop_tags)
        all_tags |= crop_tags

    return sorted(all_tags), per_crop


# === Step 4: Part List 매핑 ===
def read_part_list(xlsx_path):
    mapping = {}
    with zipfile.ZipFile(xlsx_path) as z:
        ss = []
        if 'xl/sharedStrings.xml' in z.namelist():
            tree = ET.parse(z.open('xl/sharedStrings.xml'))
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            for si in tree.findall('.//s:si', ns):
                ss.append(''.join([t.text or '' for t in si.findall('.//s:t', ns)]))
        for sheet_idx in [1, 2]:
            sp = f'xl/worksheets/sheet{sheet_idx}.xml'
            if sp not in z.namelist(): continue
            tree = ET.parse(z.open(sp))
            ns3 = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            for row in tree.findall(f'.//{{{ns3}}}row'):
                cells = {}
                for c in row:
                    v = c.find(f'{{{ns3}}}v')
                    ref = c.attrib.get('r', '')
                    col = ''.join(filter(str.isalpha, ref))
                    if v is not None and v.text:
                        cells[col] = ss[int(v.text)] if c.attrib.get('t') == 's' and int(v.text) < len(ss) else v.text
                tag = cells.get('B', '').strip().upper()
                code = cells.get('Y', cells.get('M', '')).strip()
                if tag and code and tag != '-' and tag not in ('CODE', 'NO.') and len(code) > 5:
                    mapping[tag] = code
    return mapping


# === Step 5: ERP BOM 확인 ===
def read_erp_bom(xlsx_path):
    codes = set()
    with zipfile.ZipFile(xlsx_path) as z:
        ss = []
        if 'xl/sharedStrings.xml' in z.namelist():
            tree = ET.parse(z.open('xl/sharedStrings.xml'))
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            for si in tree.findall('.//s:si', ns):
                ss.append(''.join([t.text or '' for t in si.findall('.//s:t', ns)]))
        sp = 'xl/worksheets/sheet1.xml'
        if sp in z.namelist():
            tree = ET.parse(z.open(sp))
            ns3 = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            for row in tree.findall(f'.//{{{ns3}}}row'):
                cells = {}
                for c in row:
                    v = c.find(f'{{{ns3}}}v')
                    ref = c.attrib.get('r', '')
                    col = ''.join(filter(str.isalpha, ref))
                    if v is not None and v.text:
                        cells[col] = ss[int(v.text)] if c.attrib.get('t') == 's' and int(v.text) < len(ss) else v.text
                if cells.get('A', '') == '.1':
                    code = cells.get('C', '').strip()
                    if code: codes.add(code)
    return codes


# === Step 6: 리포트 생성 ===
def generate_report(tags, part_list_map, erp_codes, output_path):
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = 'Summary'
    ws.append(['BMT GAR TAG 검증 리포트'])
    ws.append([])
    ws.append(['TAG 검출', len(tags)])
    matched = sum(1 for t in tags if part_list_map.get(t, '') in erp_codes)
    ws.append(['BOM 매칭 성공', matched])
    ws.append(['BOM 불일치/누락', len(tags) - matched])
    ws.append([])

    # Detail sheet
    ws2 = wb.create_sheet('Details')
    ws2.append(['TAG', 'Part List 코드', 'ERP BOM', '결과'])
    green = PatternFill(start_color='C6EFCE', fill_type='solid')
    red = PatternFill(start_color='FFC7CE', fill_type='solid')
    yellow = PatternFill(start_color='FFEB9C', fill_type='solid')

    for tag in sorted(tags):
        pl_code = part_list_map.get(tag, '')
        if not pl_code:
            ws2.append([tag, '(Part List에 없음)', '', '❓ 미매핑'])
            ws2[ws2.max_row][3].fill = yellow
        elif pl_code in erp_codes:
            ws2.append([tag, pl_code, '있음', '✅ 매칭'])
            ws2[ws2.max_row][3].fill = green
        else:
            ws2.append([tag, pl_code, '없음', '❌ 불일치'])
            ws2[ws2.max_row][3].fill = red

    wb.save(output_path)
    return output_path


# === Main ===
def run_pipeline(image_path, part_list_path, erp_bom_path, output_path):
    print("1/6 뷰 라벨 검출...")
    print("2/6 Min-Content 크롭...")
    crops = min_content_split(image_path)
    print(f"    {len(crops)}개 크롭")

    print("3/6 OCR + TAG 추출...")
    tags, per_crop = extract_tags_from_crops(image_path, crops)
    print(f"    {len(tags)}개 TAG: {tags}")

    print("4/6 Part List 매핑...")
    pl_map = read_part_list(part_list_path)
    print(f"    {len(pl_map)}개 매핑")

    print("5/6 ERP BOM 확인...")
    erp_codes = read_erp_bom(erp_bom_path)
    print(f"    {len(erp_codes)}개 품목")

    print("6/6 리포트 생성...")
    report = generate_report(tags, pl_map, erp_codes, output_path)
    print(f"    저장: {report}")

    # 결과 요약
    matched = sum(1 for t in tags if pl_map.get(t, '') in erp_codes)
    mismatched = [t for t in tags if pl_map.get(t, '') and pl_map[t] not in erp_codes]
    unmapped = [t for t in tags if not pl_map.get(t, '')]

    print(f"\n=== 결과 ===")
    print(f"TAG 검출: {len(tags)}개")
    print(f"BOM 매칭: {matched}건")
    print(f"BOM 불일치: {len(mismatched)}건 → {mismatched}")
    print(f"Part List 미매핑: {len(unmapped)}건 → {unmapped}")

    return {'tags': tags, 'matched': matched, 'mismatched': mismatched, 'unmapped': unmapped}


if __name__ == '__main__':
    BASE = Path(__file__).parent.parent / '도면&BOM AI검증 솔루션 관련 자료'
    run_pipeline(
        image_path='/tmp/gar-page3.png',
        part_list_path=str(BASE / '2_W5XGVU-SN2709-PLRA_Part List.xlsx'),
        erp_bom_path=str(BASE / '2_W5XHEGVU-SN2709-DNV.xlsx'),
        output_path='/tmp/bmt_report.xlsx',
    )

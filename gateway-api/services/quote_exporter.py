"""
DSE Bearing Quote Exporter Service

견적서를 Excel 및 PDF로 내보내기
"""
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class QuoteExportOptions:
    """견적서 내보내기 옵션"""
    include_logo: bool = True
    logo_path: str = ""
    include_terms: bool = True
    terms_text: str = "※ 본 견적서는 30일간 유효합니다.\n※ 부가세 별도"
    currency_format: str = "KRW"
    decimal_places: int = 0
    paper_size: str = "A4"
    orientation: str = "portrait"


class QuoteExporter:
    """견적서 내보내기 서비스"""

    def __init__(self, output_dir: Optional[str] = None):
        """
        초기화

        Args:
            output_dir: 출력 디렉토리 경로
        """
        self.output_dir = Path(output_dir) if output_dir else Path("/tmp/quotes")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_excel(
        self,
        quote_data: Dict[str, Any],
        customer_info: Optional[Dict[str, Any]] = None,
        options: Optional[QuoteExportOptions] = None
    ) -> bytes:
        """
        견적서를 Excel로 내보내기

        Args:
            quote_data: 견적 데이터
            customer_info: 고객 정보
            options: 내보내기 옵션

        Returns:
            Excel 파일 바이트
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl 모듈이 필요합니다: pip install openpyxl")
            raise

        options = options or QuoteExportOptions()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "견적서"

        # 스타일 정의
        title_font = Font(name="맑은 고딕", size=18, bold=True)
        header_font = Font(name="맑은 고딕", size=11, bold=True)
        normal_font = Font(name="맑은 고딕", size=10)
        money_font = Font(name="맑은 고딕", size=10)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = "견 적 서"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

        # 견적 정보
        row = 3
        ws[f'A{row}'] = "견적번호:"
        ws[f'B{row}'] = quote_data.get('quote_number', '')
        ws[f'D{row}'] = "견적일자:"
        ws[f'E{row}'] = quote_data.get('date', datetime.now().strftime('%Y-%m-%d'))

        row += 1
        if customer_info:
            ws[f'A{row}'] = "고객사:"
            ws[f'B{row}'] = customer_info.get('customer_name', '')
            ws[f'D{row}'] = "담당자:"
            ws[f'E{row}'] = customer_info.get('contact_name', '')

        # 열 헤더
        row += 2
        headers = ['No.', '품명/규격', '재질', '수량', '단가', '금액', '비고']
        col_widths = [6, 35, 12, 8, 15, 18, 15]

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[row].height = 25

        # 항목 데이터
        start_row = row + 1
        for item in quote_data.get('line_items', quote_data.get('items', [])):
            row += 1
            item_no = item.get('no', '')
            desc = item.get('description', '')
            material = item.get('material', '')
            qty = item.get('qty', item.get('quantity', 1))
            unit_price = item.get('unit_price', 0)
            total_price = item.get('total_price', 0)
            remark = item.get('remark', '')

            values = [item_no, desc, material, qty, unit_price, total_price, remark]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border

                if col in [4, 5, 6]:  # 수량, 단가, 금액
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if col in [5, 6] and isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal='center' if col == 1 else 'left', vertical='center')

        # 합계
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "소계"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        ws[f'A{row}'].border = thin_border

        ws.merge_cells(f'E{row}:F{row}')
        subtotal = quote_data.get('subtotal', 0)
        ws[f'E{row}'] = subtotal
        ws[f'E{row}'].font = money_font
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = thin_border

        # 할인
        discount = quote_data.get('discount_amount', quote_data.get('discount', 0))
        if discount > 0:
            row += 1
            ws.merge_cells(f'A{row}:D{row}')
            discount_rate = quote_data.get('quantity_discount_rate', 0) * 100
            ws[f'A{row}'] = f"수량 할인 ({discount_rate:.0f}%)"
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = Alignment(horizontal='right')
            ws[f'A{row}'].border = thin_border

            ws.merge_cells(f'E{row}:F{row}')
            ws[f'E{row}'] = -discount
            ws[f'E{row}'].font = money_font
            ws[f'E{row}'].number_format = '#,##0'
            ws[f'E{row}'].alignment = Alignment(horizontal='right')
            ws[f'E{row}'].border = thin_border

        # 부가세
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        tax_rate = quote_data.get('tax_rate', 0.1) * 100
        ws[f'A{row}'] = f"부가세 ({tax_rate:.0f}%)"
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        ws[f'A{row}'].border = thin_border

        ws.merge_cells(f'E{row}:F{row}')
        tax = quote_data.get('tax', 0)
        ws[f'E{row}'] = tax
        ws[f'E{row}'].font = money_font
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = thin_border

        # 총액
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "총액"
        ws[f'A{row}'].font = Font(name="맑은 고딕", size=12, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].fill = total_fill

        ws.merge_cells(f'E{row}:F{row}')
        total = quote_data.get('total', 0)
        ws[f'E{row}'] = total
        ws[f'E{row}'].font = Font(name="맑은 고딕", size=12, bold=True)
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].fill = total_fill

        currency = quote_data.get('currency', 'KRW')
        ws[f'G{row}'] = currency
        ws[f'G{row}'].font = header_font
        ws[f'G{row}'].alignment = Alignment(horizontal='center')

        # 비고/조건
        if options.include_terms:
            row += 3
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = options.terms_text
            ws[f'A{row}'].font = Font(name="맑은 고딕", size=9, italic=True)
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)

        # 바이트로 반환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_to_pdf(
        self,
        quote_data: Dict[str, Any],
        customer_info: Optional[Dict[str, Any]] = None,
        options: Optional[QuoteExportOptions] = None
    ) -> bytes:
        """
        견적서를 PDF로 내보내기

        Args:
            quote_data: 견적 데이터
            customer_info: 고객 정보
            options: 내보내기 옵션

        Returns:
            PDF 파일 바이트
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            logger.error("reportlab 모듈이 필요합니다: pip install reportlab")
            raise

        options = options or QuoteExportOptions()

        # 한글 폰트 등록 (시스템에 있는 폰트 사용)
        font_paths = [
            "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
            "/usr/share/fonts/nanum/NanumGothic.ttf",
            "/System/Library/Fonts/AppleGothic.ttf",
            "C:\\Windows\\Fonts\\malgun.ttf",
        ]

        font_registered = False
        for font_path in font_paths:
            if Path(font_path).exists():
                try:
                    pdfmetrics.registerFont(TTFont('Korean', font_path))
                    font_registered = True
                    break
                except Exception:
                    continue

        if not font_registered:
            logger.warning("한글 폰트를 찾을 수 없습니다. 기본 폰트 사용")
            korean_font = 'Helvetica'
        else:
            korean_font = 'Korean'

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )

        elements = []
        styles = getSampleStyleSheet()

        # 제목 스타일
        title_style = ParagraphStyle(
            'TitleKorean',
            parent=styles['Title'],
            fontName=korean_font,
            fontSize=24,
            alignment=1,  # Center
            spaceAfter=20
        )

        normal_style = ParagraphStyle(
            'NormalKorean',
            parent=styles['Normal'],
            fontName=korean_font,
            fontSize=10,
        )

        # 제목
        elements.append(Paragraph("견 적 서", title_style))
        elements.append(Spacer(1, 10*mm))

        # 견적 정보
        quote_number = quote_data.get('quote_number', '')
        quote_date = quote_data.get('date', datetime.now().strftime('%Y-%m-%d'))
        info_data = [
            [f"견적번호: {quote_number}", f"견적일자: {quote_date}"]
        ]
        if customer_info:
            info_data.append([
                f"고객사: {customer_info.get('customer_name', '')}",
                f"담당자: {customer_info.get('contact_name', '')}"
            ])

        info_table = Table(info_data, colWidths=[90*mm, 80*mm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), korean_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10*mm))

        # 항목 테이블
        header = ['No.', '품명/규격', '재질', '수량', '단가', '금액']
        table_data = [header]

        for item in quote_data.get('line_items', quote_data.get('items', [])):
            row = [
                str(item.get('no', '')),
                str(item.get('description', '')),
                str(item.get('material', '')),
                str(item.get('qty', item.get('quantity', 1))),
                f"{item.get('unit_price', 0):,.0f}",
                f"{item.get('total_price', 0):,.0f}",
            ]
            table_data.append(row)

        # 합계 행
        subtotal = quote_data.get('subtotal', 0)
        table_data.append(['', '', '', '', '소계', f"{subtotal:,.0f}"])

        discount = quote_data.get('discount_amount', quote_data.get('discount', 0))
        if discount > 0:
            table_data.append(['', '', '', '', '할인', f"-{discount:,.0f}"])

        tax = quote_data.get('tax', 0)
        table_data.append(['', '', '', '', '부가세', f"{tax:,.0f}"])

        total = quote_data.get('total', 0)
        currency = quote_data.get('currency', 'KRW')
        table_data.append(['', '', '', '', '총액', f"{total:,.0f} {currency}"])

        col_widths = [15*mm, 60*mm, 25*mm, 15*mm, 25*mm, 30*mm]
        main_table = Table(table_data, colWidths=col_widths)

        main_table.setStyle(TableStyle([
            # 헤더
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), korean_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            # 데이터
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            # 총액 행 강조
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightyellow),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
        ]))
        elements.append(main_table)

        # 비고
        if options.include_terms:
            elements.append(Spacer(1, 15*mm))
            elements.append(Paragraph(options.terms_text.replace('\n', '<br/>'), normal_style))

        doc.build(elements)
        output.seek(0)
        return output.getvalue()

    def save_quote(
        self,
        quote_data: Dict[str, Any],
        format: str = "excel",
        filename: Optional[str] = None,
        customer_info: Optional[Dict[str, Any]] = None,
        options: Optional[QuoteExportOptions] = None
    ) -> str:
        """
        견적서 파일 저장

        Args:
            quote_data: 견적 데이터
            format: 출력 형식 (excel, pdf)
            filename: 파일명 (없으면 자동 생성)
            customer_info: 고객 정보
            options: 내보내기 옵션

        Returns:
            저장된 파일 경로
        """
        if not filename:
            quote_number = quote_data.get('quote_number', datetime.now().strftime('%Y%m%d%H%M%S'))
            ext = 'xlsx' if format == 'excel' else 'pdf'
            filename = f"quote_{quote_number}.{ext}"

        filepath = self.output_dir / filename

        if format == 'excel':
            content = self.export_to_excel(quote_data, customer_info, options)
        elif format == 'pdf':
            content = self.export_to_pdf(quote_data, customer_info, options)
        else:
            raise ValueError(f"지원하지 않는 형식: {format}")

        with open(filepath, 'wb') as f:
            f.write(content)

        logger.info(f"견적서 저장: {filepath}")
        return str(filepath)


# 싱글톤 인스턴스
_exporter_instance = None


def get_quote_exporter(output_dir: Optional[str] = None) -> QuoteExporter:
    """견적서 내보내기 서비스 인스턴스 반환"""
    global _exporter_instance
    if _exporter_instance is None:
        _exporter_instance = QuoteExporter(output_dir)
    return _exporter_instance

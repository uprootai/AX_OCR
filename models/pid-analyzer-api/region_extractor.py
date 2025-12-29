"""
Region Text Extractor Module
영역 기반 텍스트 추출 및 규칙 엔진

핵심 기능:
- Line Detector 영역 + OCR 텍스트 매칭
- 유연한 규칙 기반 추출 (YAML 설정)
- UI에서 규칙 편집 가능

저자: Claude AI (Opus 4.5)
생성일: 2025-12-29
"""
import os
import re
import yaml
import logging
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field, asdict
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# =====================
# Data Classes
# =====================

@dataclass
class ExtractionPattern:
    """추출 패턴 정의"""
    type: str  # valve_tag, equipment_tag, label, etc.
    regex: str
    description: str = ""
    priority: int = 0

    def match(self, text: str) -> Optional[Dict[str, Any]]:
        """텍스트에서 패턴 매칭"""
        try:
            pattern = re.compile(self.regex, re.IGNORECASE)
            match = pattern.search(text)
            if match:
                return {
                    "matched_text": match.group(0),
                    "groups": match.groups() if match.groups() else [],
                    "type": self.type,
                    "pattern": self.regex
                }
        except re.error as e:
            logger.warning(f"Regex error in pattern '{self.regex}': {e}")
        return None


@dataclass
class RegionTextPattern:
    """영역 식별용 텍스트 패턴"""
    pattern: str
    case_insensitive: bool = True
    is_regex: bool = False

    def matches(self, text: str) -> bool:
        """텍스트가 패턴과 매칭되는지 확인"""
        try:
            if self.is_regex:
                flags = re.IGNORECASE if self.case_insensitive else 0
                return bool(re.search(self.pattern, text, flags))
            else:
                compare_text = text.lower() if self.case_insensitive else text
                compare_pattern = self.pattern.lower() if self.case_insensitive else self.pattern
                return compare_pattern in compare_text
        except Exception as e:
            logger.warning(f"Pattern match error: {e}")
            return False


@dataclass
class RegionCriteria:
    """영역 매칭 기준"""
    line_styles: List[str] = field(default_factory=lambda: ["dashed", "dash_dot"])
    min_area: int = 1000
    max_area: Optional[int] = None
    aspect_ratio_min: Optional[float] = None
    aspect_ratio_max: Optional[float] = None


@dataclass
class ExtractionRule:
    """추출 규칙 정의"""
    id: str
    name: str
    description: str = ""
    enabled: bool = True

    # 영역 매칭 기준
    region_criteria: RegionCriteria = field(default_factory=RegionCriteria)

    # 영역 식별용 텍스트 패턴 (이 텍스트가 있어야 해당 영역으로 인식)
    region_text_patterns: List[RegionTextPattern] = field(default_factory=list)

    # 추출할 패턴들
    extraction_patterns: List[ExtractionPattern] = field(default_factory=list)

    # 출력 설정
    output_type: str = "list"  # list, excel, json
    output_filename_template: str = "{rule_id}_{timestamp}"

    # 메타데이터 (UI용)
    category: str = "general"
    icon: str = "list"
    color: str = "#3b82f6"

    def to_dict(self) -> Dict[str, Any]:
        """딕셔너리로 변환 (API 응답용)"""
        return {
            "id": self.id,
            "name": self.name,
            "description": self.description,
            "enabled": self.enabled,
            "region_criteria": {
                "line_styles": self.region_criteria.line_styles,
                "min_area": self.region_criteria.min_area,
                "max_area": self.region_criteria.max_area,
                "aspect_ratio_min": self.region_criteria.aspect_ratio_min,
                "aspect_ratio_max": self.region_criteria.aspect_ratio_max,
            },
            "region_text_patterns": [
                {"pattern": p.pattern, "case_insensitive": p.case_insensitive, "is_regex": p.is_regex}
                for p in self.region_text_patterns
            ],
            "extraction_patterns": [
                {"type": p.type, "regex": p.regex, "description": p.description, "priority": p.priority}
                for p in self.extraction_patterns
            ],
            "output_type": self.output_type,
            "output_filename_template": self.output_filename_template,
            "category": self.category,
            "icon": self.icon,
            "color": self.color,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "ExtractionRule":
        """딕셔너리에서 생성"""
        region_criteria_data = data.get("region_criteria", {})
        region_criteria = RegionCriteria(
            line_styles=region_criteria_data.get("line_styles", ["dashed", "dash_dot"]),
            min_area=region_criteria_data.get("min_area", 1000),
            max_area=region_criteria_data.get("max_area"),
            aspect_ratio_min=region_criteria_data.get("aspect_ratio_min"),
            aspect_ratio_max=region_criteria_data.get("aspect_ratio_max"),
        )

        region_text_patterns = [
            RegionTextPattern(
                pattern=p.get("pattern", ""),
                case_insensitive=p.get("case_insensitive", True),
                is_regex=p.get("is_regex", False)
            )
            for p in data.get("region_text_patterns", [])
        ]

        extraction_patterns = [
            ExtractionPattern(
                type=p.get("type", "unknown"),
                regex=p.get("regex", ""),
                description=p.get("description", ""),
                priority=p.get("priority", 0)
            )
            for p in data.get("extraction_patterns", [])
        ]

        return cls(
            id=data.get("id", "unknown"),
            name=data.get("name", "Unknown Rule"),
            description=data.get("description", ""),
            enabled=data.get("enabled", True),
            region_criteria=region_criteria,
            region_text_patterns=region_text_patterns,
            extraction_patterns=extraction_patterns,
            output_type=data.get("output_type", "list"),
            output_filename_template=data.get("output_filename_template", "{rule_id}_{timestamp}"),
            category=data.get("category", "general"),
            icon=data.get("icon", "list"),
            color=data.get("color", "#3b82f6"),
        )


@dataclass
class MatchedRegion:
    """매칭된 영역 정보"""
    region_id: int
    rule_id: str
    bbox: List[float]
    region_texts: List[str]  # 영역 식별에 사용된 텍스트
    extracted_items: List[Dict[str, Any]]  # 추출된 항목들
    confidence: float = 1.0


# =====================
# Rule Manager
# =====================

class RuleManager:
    """규칙 관리자 - YAML 파일 기반 규칙 저장/로드"""

    def __init__(self, rules_file: str = None):
        if rules_file is None:
            # 기본 경로: 현재 디렉토리의 region_rules.yaml
            rules_file = os.path.join(os.path.dirname(__file__), "region_rules.yaml")

        self.rules_file = Path(rules_file)
        self.rules: Dict[str, ExtractionRule] = {}
        self._load_rules()

    def _load_rules(self):
        """YAML 파일에서 규칙 로드"""
        if not self.rules_file.exists():
            logger.info(f"Rules file not found: {self.rules_file}. Using defaults.")
            self._create_default_rules()
            return

        try:
            with open(self.rules_file, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)

            if data and "rules" in data:
                for rule_data in data["rules"]:
                    rule = ExtractionRule.from_dict(rule_data)
                    self.rules[rule.id] = rule

            logger.info(f"Loaded {len(self.rules)} rules from {self.rules_file}")

        except Exception as e:
            logger.error(f"Error loading rules: {e}")
            self._create_default_rules()

    def _create_default_rules(self):
        """기본 규칙 생성"""
        # BWMS Valve Signal List 규칙
        bwms_valve_rule = ExtractionRule(
            id="valve_signal_bwms",
            name="Valve Signal List (BWMS)",
            description="SIGNAL FOR BWMS 영역 내 밸브 ID 추출",
            enabled=True,
            region_criteria=RegionCriteria(
                line_styles=["dashed", "dash_dot"],
                min_area=1000
            ),
            region_text_patterns=[
                RegionTextPattern(pattern="SIGNAL.*FOR.*BWMS", is_regex=True),
                RegionTextPattern(pattern="SIGNAL FOR BWMS"),
                RegionTextPattern(pattern="SIGNAL_FOR_BWMS"),
            ],
            extraction_patterns=[
                ExtractionPattern(
                    type="valve_tag",
                    regex=r"^(BWV|BAV|BCV|BBV|BXV|BSV|BFV|BRV)\d+[A-Z]?$",
                    description="BWMS 밸브 태그 (BWV, BAV, BCV 등)",
                    priority=10
                ),
                ExtractionPattern(
                    type="valve_tag",
                    regex=r"^(FCV|TCV|PCV|LCV|HCV|SCV|ACV)\d+[A-Z]?$",
                    description="일반 제어 밸브 태그",
                    priority=5
                ),
                ExtractionPattern(
                    type="valve_tag",
                    regex=r"^[A-Z]{2,4}-?\d{1,4}[A-Z]?$",
                    description="일반 밸브 태그 패턴",
                    priority=1
                ),
            ],
            output_type="excel",
            output_filename_template="Valve_Signal_List_{timestamp}",
            category="bwms",
            icon="valve",
            color="#ef4444"
        )

        # 일반 Signal 영역 규칙
        general_signal_rule = ExtractionRule(
            id="signal_region_general",
            name="Signal Region (General)",
            description="일반 SIGNAL 영역 내 태그 추출",
            enabled=True,
            region_criteria=RegionCriteria(
                line_styles=["dashed", "dash_dot", "dotted"],
                min_area=500
            ),
            region_text_patterns=[
                RegionTextPattern(pattern="SIGNAL", is_regex=False),
            ],
            extraction_patterns=[
                ExtractionPattern(
                    type="tag",
                    regex=r"^[A-Z]{1,5}-?\d{1,4}[A-Z]?$",
                    description="일반 태그 패턴",
                    priority=1
                ),
            ],
            output_type="list",
            category="general",
            icon="tag",
            color="#3b82f6"
        )

        # 알람 영역 규칙 (예시)
        alarm_rule = ExtractionRule(
            id="alarm_bypass",
            name="Alarm By-pass",
            description="ALARM BY-PASS 영역 내 밸브 추출",
            enabled=True,
            region_criteria=RegionCriteria(
                line_styles=["dashed", "dash_dot"],
                min_area=800
            ),
            region_text_patterns=[
                RegionTextPattern(pattern="ALARM.*BY-?PASS", is_regex=True),
                RegionTextPattern(pattern="ALARM BYPASS"),
            ],
            extraction_patterns=[
                ExtractionPattern(
                    type="valve_tag",
                    regex=r"^[A-Z]{2,4}-?\d{1,4}$",
                    description="알람 바이패스 밸브",
                    priority=5
                ),
            ],
            output_type="list",
            category="bwms",
            icon="bell",
            color="#f59e0b"
        )

        self.rules = {
            bwms_valve_rule.id: bwms_valve_rule,
            general_signal_rule.id: general_signal_rule,
            alarm_rule.id: alarm_rule,
        }

        # 기본 규칙 저장
        self.save_rules()

    def save_rules(self):
        """규칙을 YAML 파일로 저장"""
        try:
            data = {
                "version": "1.0",
                "updated_at": datetime.now().isoformat(),
                "rules": [rule.to_dict() for rule in self.rules.values()]
            }

            with open(self.rules_file, 'w', encoding='utf-8') as f:
                yaml.dump(data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

            logger.info(f"Saved {len(self.rules)} rules to {self.rules_file}")

        except Exception as e:
            logger.error(f"Error saving rules: {e}")
            raise

    def get_rule(self, rule_id: str) -> Optional[ExtractionRule]:
        """규칙 조회"""
        return self.rules.get(rule_id)

    def get_all_rules(self) -> List[ExtractionRule]:
        """모든 규칙 조회"""
        return list(self.rules.values())

    def get_enabled_rules(self) -> List[ExtractionRule]:
        """활성화된 규칙만 조회"""
        return [r for r in self.rules.values() if r.enabled]

    def add_rule(self, rule: ExtractionRule) -> bool:
        """규칙 추가"""
        if rule.id in self.rules:
            logger.warning(f"Rule '{rule.id}' already exists")
            return False

        self.rules[rule.id] = rule
        self.save_rules()
        return True

    def update_rule(self, rule_id: str, rule_data: Dict[str, Any]) -> Optional[ExtractionRule]:
        """규칙 업데이트"""
        if rule_id not in self.rules:
            logger.warning(f"Rule '{rule_id}' not found")
            return None

        # ID는 변경 불가
        rule_data["id"] = rule_id

        updated_rule = ExtractionRule.from_dict(rule_data)
        self.rules[rule_id] = updated_rule
        self.save_rules()

        return updated_rule

    def delete_rule(self, rule_id: str) -> bool:
        """규칙 삭제"""
        if rule_id not in self.rules:
            return False

        del self.rules[rule_id]
        self.save_rules()
        return True

    def get_rules_by_category(self, category: str) -> List[ExtractionRule]:
        """카테고리별 규칙 조회"""
        return [r for r in self.rules.values() if r.category == category]


# =====================
# Region Text Extractor
# =====================

class RegionTextExtractor:
    """영역-텍스트 매칭 및 추출 엔진"""

    def __init__(self, rule_manager: RuleManager = None):
        self.rule_manager = rule_manager or RuleManager()

    def _bbox_contains_point(self, bbox: List[float], point: Tuple[float, float], margin: float = 10) -> bool:
        """bbox가 점을 포함하는지 확인 (마진 포함)"""
        if len(bbox) < 4:
            return False

        x1, y1, x2, y2 = bbox[:4]
        px, py = point

        return (x1 - margin) <= px <= (x2 + margin) and (y1 - margin) <= py <= (y2 + margin)

    def _get_text_center(self, text_item: Dict[str, Any]) -> Optional[Tuple[float, float]]:
        """OCR 텍스트 아이템의 중심점 계산"""
        # PaddleOCR 형식: bbox가 [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
        bbox = text_item.get("bbox", text_item.get("box", []))

        if not bbox:
            return None

        if isinstance(bbox[0], list):
            # [[x1,y1],[x2,y2],[x3,y3],[x4,y4]] 형식
            xs = [p[0] for p in bbox]
            ys = [p[1] for p in bbox]
            return ((min(xs) + max(xs)) / 2, (min(ys) + max(ys)) / 2)
        elif len(bbox) >= 4:
            # [x1, y1, x2, y2] 형식
            return ((bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2)

        return None

    def _calculate_region_area(self, region: Dict[str, Any]) -> float:
        """영역 면적 계산"""
        bbox = region.get("bbox", region.get("bounding_box", []))
        if len(bbox) < 4:
            return 0

        x1, y1, x2, y2 = bbox[:4]
        return abs(x2 - x1) * abs(y2 - y1)

    def _region_matches_criteria(self, region: Dict[str, Any], criteria: RegionCriteria) -> bool:
        """영역이 기준과 매칭되는지 확인"""
        # 라인 스타일 확인 - 여러 필드명 지원
        region_style = region.get("line_style") or region.get("enclosing_line_style") or region.get("style", "unknown")

        # "mixed" 스타일은 점선 영역으로 간주 (dashed, dash_dot 포함 시)
        if region_style == "mixed":
            # lines_inside에서 dashed/dash_dot 라인이 있는지 확인
            lines_inside = region.get("lines_inside", [])
            has_dashed = any(
                line.get("line_style") in criteria.line_styles
                for line in lines_inside
            )
            if has_dashed:
                region_style = "dashed"  # mixed이지만 dashed 라인 포함

        if region_style not in criteria.line_styles and region_style != "mixed":
            return False

        # 면적 확인
        area = self._calculate_region_area(region)
        if area < criteria.min_area:
            return False

        if criteria.max_area and area > criteria.max_area:
            return False

        # 종횡비 확인
        bbox = region.get("bbox", region.get("bounding_box", []))
        if len(bbox) >= 4:
            width = abs(bbox[2] - bbox[0])
            height = abs(bbox[3] - bbox[1])
            if height > 0:
                aspect_ratio = width / height

                if criteria.aspect_ratio_min and aspect_ratio < criteria.aspect_ratio_min:
                    return False
                if criteria.aspect_ratio_max and aspect_ratio > criteria.aspect_ratio_max:
                    return False

        return True

    def _find_texts_in_region(self, region_bbox: List[float], texts: List[Dict], margin: float = 20) -> List[Dict]:
        """영역 내에 있는 텍스트 찾기"""
        texts_in_region = []

        for text_item in texts:
            center = self._get_text_center(text_item)
            if center and self._bbox_contains_point(region_bbox, center, margin):
                texts_in_region.append(text_item)

        return texts_in_region

    def _region_matches_text_patterns(
        self,
        region_texts: List[Dict],
        patterns: List[RegionTextPattern]
    ) -> Tuple[bool, List[str]]:
        """영역 내 텍스트가 패턴과 매칭되는지 확인

        개선된 매칭 로직:
        1. 개별 텍스트가 패턴과 매칭되는지 확인
        2. 영역 내 모든 텍스트를 결합하여 패턴과 매칭 (OCR이 텍스트를 분리할 경우 대응)
        """
        if not patterns:
            return True, []  # 패턴이 없으면 모든 영역 매칭

        matched_texts = []

        # 1단계: 개별 텍스트 매칭 확인
        for text_item in region_texts:
            text = text_item.get("text", "")

            for pattern in patterns:
                if pattern.matches(text):
                    matched_texts.append(text)
                    break

        # 1단계에서 매칭되면 바로 반환
        if matched_texts:
            return True, matched_texts

        # 2단계: 영역 내 모든 텍스트를 결합하여 매칭 시도
        # OCR이 "SIGNAL"과 "FOR BWMS"를 분리해서 반환하는 경우 대응
        all_texts = [item.get("text", "") for item in region_texts]
        combined_text = " ".join(all_texts)
        combined_text_nospace = "".join(all_texts)  # 공백 없는 버전도 시도

        for pattern in patterns:
            if pattern.matches(combined_text) or pattern.matches(combined_text_nospace):
                # 매칭 성공 시, 패턴에 포함된 키워드를 가진 텍스트만 반환
                # 예: "SIGNAL FOR BWMS" 패턴 -> "SIGNAL"과 "FOR BWMS" 텍스트만 반환
                pattern_keywords = pattern.pattern.replace("_", " ").replace(".*", " ").split()
                relevant_texts = []
                for text in all_texts:
                    text_check = text.upper() if pattern.case_insensitive else text
                    for keyword in pattern_keywords:
                        keyword_check = keyword.upper() if pattern.case_insensitive else keyword
                        if keyword_check in text_check:
                            if text not in relevant_texts:
                                relevant_texts.append(text)
                            break
                if relevant_texts:
                    return True, relevant_texts
                # 키워드 매칭 실패 시 첫 번째 텍스트만 반환
                return True, [all_texts[0]] if all_texts else []

        # 3단계: 개별 패턴 키워드들이 각각 존재하는지 확인
        # 예: "SIGNAL FOR BWMS" -> "SIGNAL", "FOR", "BWMS" 각각 존재 확인
        # 단, 실제 매칭된 텍스트만 반환 (밸브 태그 등은 제외)
        for pattern in patterns:
            if not pattern.is_regex:
                # 단순 문자열 패턴의 경우, 키워드 분리
                keywords = pattern.pattern.replace("_", " ").split()
                found_all = True
                found_keywords = []  # 실제 키워드가 포함된 텍스트만 저장

                for keyword in keywords:
                    keyword_upper = keyword.upper() if pattern.case_insensitive else keyword
                    found = False
                    for text in all_texts:
                        text_check = text.upper() if pattern.case_insensitive else text
                        # 키워드가 텍스트에 포함되어 있는지 확인
                        if keyword_upper in text_check:
                            found = True
                            # 해당 텍스트가 키워드를 포함하면 추가
                            if text not in found_keywords:
                                found_keywords.append(text)
                            break
                    if not found:
                        found_all = False
                        break

                if found_all and found_keywords:
                    # 찾은 키워드 텍스트만 반환 (밸브 태그 등 다른 텍스트는 제외)
                    return True, found_keywords

        # 최소 1개의 패턴이 매칭되어야 함
        return len(matched_texts) > 0, matched_texts

    def _extract_from_texts(
        self,
        texts: List[Dict],
        patterns: List[ExtractionPattern],
        exclude_texts: List[str] = None
    ) -> List[Dict[str, Any]]:
        """텍스트에서 패턴에 매칭되는 항목 추출"""
        if exclude_texts is None:
            exclude_texts = []

        extracted = []
        seen = set()  # 중복 방지

        # 우선순위 순으로 패턴 정렬
        sorted_patterns = sorted(patterns, key=lambda p: p.priority, reverse=True)

        for text_item in texts:
            text = text_item.get("text", "").strip()

            # 영역 식별 텍스트는 제외
            if any(exc.lower() in text.lower() for exc in exclude_texts):
                continue

            # 너무 긴 텍스트는 제외 (일반적으로 태그는 짧음)
            if len(text) > 20:
                continue

            for pattern in sorted_patterns:
                match_result = pattern.match(text)
                if match_result:
                    matched_text = match_result["matched_text"]

                    if matched_text not in seen:
                        seen.add(matched_text)

                        # 텍스트 위치 정보 추출
                        center = self._get_text_center(text_item)
                        bbox = text_item.get("bbox", text_item.get("box", []))

                        extracted.append({
                            "id": matched_text,
                            "type": pattern.type,
                            "matched_text": matched_text,
                            "original_text": text,
                            "confidence": text_item.get("confidence", 1.0),
                            "center": center,
                            "bbox": bbox,
                            "pattern_description": pattern.description,
                        })
                    break  # 한 텍스트당 하나의 패턴만 매칭

        return extracted

    def extract(
        self,
        regions: List[Dict],
        texts: List[Dict],
        rule_ids: List[str] = None,
        text_margin: float = 30
    ) -> Dict[str, Any]:
        """
        영역에서 규칙에 따라 텍스트 추출

        Args:
            regions: Line Detector에서 검출된 영역 목록
            texts: PaddleOCR에서 검출된 텍스트 목록
            rule_ids: 적용할 규칙 ID 목록 (None이면 활성화된 모든 규칙)
            text_margin: 텍스트 매칭 시 마진 (픽셀)

        Returns:
            추출 결과 (규칙별 매칭 영역 및 추출 항목)
        """
        # 적용할 규칙 결정
        if rule_ids:
            rules = [self.rule_manager.get_rule(rid) for rid in rule_ids]
            rules = [r for r in rules if r is not None and r.enabled]
        else:
            rules = self.rule_manager.get_enabled_rules()

        if not rules:
            return {
                "success": True,
                "message": "No enabled rules found",
                "results": [],
                "statistics": {}
            }

        results_by_rule = {}

        for rule in rules:
            matched_regions = []
            logger.info(f"Processing rule: {rule.id}")
            criteria_passed = 0
            texts_found = 0
            pattern_matched = 0

            for region_idx, region in enumerate(regions):
                # 1. 영역이 기준에 맞는지 확인
                if not self._region_matches_criteria(region, rule.region_criteria):
                    continue
                criteria_passed += 1

                # 2. 영역 내 텍스트 찾기
                region_bbox = region.get("bbox", region.get("bounding_box", []))
                if len(region_bbox) < 4:
                    continue

                texts_in_region = self._find_texts_in_region(region_bbox, texts, text_margin)

                if not texts_in_region:
                    continue
                texts_found += 1
                text_contents = [t.get("text", "") for t in texts_in_region]
                logger.debug(f"Region {region_idx} texts: {text_contents[:5]}")

                # 3. 텍스트 패턴 매칭 확인
                matches, matched_pattern_texts = self._region_matches_text_patterns(
                    texts_in_region,
                    rule.region_text_patterns
                )

                if not matches:
                    continue
                pattern_matched += 1
                logger.info(f"Region {region_idx} matched patterns: {matched_pattern_texts}")

                # 4. 추출 패턴 적용
                extracted_items = self._extract_from_texts(
                    texts_in_region,
                    rule.extraction_patterns,
                    exclude_texts=matched_pattern_texts  # 영역 식별 텍스트 제외
                )

                if extracted_items:
                    matched_regions.append(MatchedRegion(
                        region_id=region_idx,
                        rule_id=rule.id,
                        bbox=region_bbox,
                        region_texts=matched_pattern_texts,
                        extracted_items=extracted_items,
                        confidence=sum(i["confidence"] for i in extracted_items) / len(extracted_items)
                    ))
                    logger.info(f"Region {region_idx} extracted: {[i['id'] for i in extracted_items]}")

            logger.info(f"Rule {rule.id}: criteria_passed={criteria_passed}, texts_found={texts_found}, pattern_matched={pattern_matched}")

            if matched_regions:
                results_by_rule[rule.id] = {
                    "rule": rule.to_dict(),
                    "matched_regions": [
                        {
                            "region_id": mr.region_id,
                            "bbox": mr.bbox,
                            "region_texts": mr.region_texts,
                            "extracted_items": mr.extracted_items,
                            "confidence": mr.confidence
                        }
                        for mr in matched_regions
                    ],
                    "total_extracted": sum(len(mr.extracted_items) for mr in matched_regions)
                }

        # 통계 생성
        all_extracted = []
        for rule_result in results_by_rule.values():
            for region in rule_result["matched_regions"]:
                all_extracted.extend(region["extracted_items"])

        # 중복 제거 후 통합 리스트
        unique_items = {}
        for item in all_extracted:
            key = item["id"]
            if key not in unique_items:
                unique_items[key] = item

        return {
            "success": True,
            "results_by_rule": results_by_rule,
            "all_extracted_items": list(unique_items.values()),
            "statistics": {
                "total_regions": len(regions),
                "total_texts": len(texts),
                "rules_applied": len(rules),
                "rules_matched": len(results_by_rule),
                "total_matched_regions": sum(
                    len(r["matched_regions"]) for r in results_by_rule.values()
                ),
                "total_extracted_items": len(unique_items)
            }
        }


# =====================
# Excel Export
# =====================

def generate_valve_signal_excel(
    extracted_items: List[Dict[str, Any]],
    project_info: Dict[str, str] = None,
    rule_name: str = "Valve Signal List"
) -> bytes:
    """
    밸브 시그널 리스트 Excel 생성

    Args:
        extracted_items: 추출된 밸브 항목 목록
        project_info: 프로젝트 정보 (name, drawing_no, date 등)
        rule_name: 규칙 이름

    Returns:
        Excel 파일 바이트 데이터
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Valve Signal List"

    # 스타일 정의
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 제목 행
    current_row = 1
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = rule_name
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws[f'A{current_row}'].alignment = center_align
    current_row += 1

    # 프로젝트 정보
    if project_info:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        info_text = f"Project: {project_info.get('name', 'N/A')} | Drawing: {project_info.get('drawing_no', 'N/A')} | Date: {project_info.get('date', datetime.now().strftime('%Y-%m-%d'))}"
        ws[f'A{current_row}'] = info_text
        ws[f'A{current_row}'].font = Font(size=10, italic=True)
        current_row += 2
    else:
        current_row += 1

    # 헤더
    headers = ["No.", "Valve ID", "Type", "Category", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    current_row += 1

    # 데이터
    for idx, item in enumerate(extracted_items, 1):
        valve_id = item.get("id", item.get("matched_text", ""))
        item_type = item.get("type", "unknown")

        # 밸브 카테고리 추론
        category = "Unknown"
        if valve_id.startswith(("BWV", "BAV", "BCV", "BBV", "BXV", "BSV", "BFV", "BRV")):
            category = "BWMS Valve"
        elif valve_id.startswith(("FCV", "TCV", "PCV", "LCV")):
            category = "Control Valve"
        elif valve_id.startswith(("HCV", "SCV", "ACV")):
            category = "Special Valve"

        row_data = [
            idx,
            valve_id,
            item_type.replace("_", " ").title(),
            category,
            f"{item.get('confidence', 1.0):.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align

        current_row += 1

    # 컬럼 너비 조정
    col_widths = [8, 15, 15, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 푸터
    current_row += 1
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = f"Generated by P&ID Analyzer | Total: {len(extracted_items)} items"
    ws[f'A{current_row}'].font = Font(size=9, italic=True, color="666666")

    # 바이트로 변환
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


# =====================
# Singleton Instance
# =====================

_rule_manager_instance = None

def get_rule_manager() -> RuleManager:
    """RuleManager 싱글턴 인스턴스 반환"""
    global _rule_manager_instance
    if _rule_manager_instance is None:
        _rule_manager_instance = RuleManager()
    return _rule_manager_instance


def get_extractor() -> RegionTextExtractor:
    """RegionTextExtractor 인스턴스 반환"""
    return RegionTextExtractor(get_rule_manager())

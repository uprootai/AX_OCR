"""
Equipment Analyzer - 산업별 장비 태그 분석 모듈

다양한 산업 분야의 P&ID 도면에서 장비 태그를 인식하는 범용 모듈.
프로파일 시스템을 통해 여러 산업/업체에 맞춤 설정 적용 가능.

지원 프로파일:
- bwms: 선박 평형수 처리 시스템 (BWMS) - ECU, FMU, HGU 등
- hvac: 공조 시스템 - AHU, FCU, VAV 등 (예정)
- process: 일반 공정 - PUMP, TANK, VALVE 등 (예정)

포트: 5018 (PID Analyzer API 내 통합)
"""

import re
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime

logger = logging.getLogger(__name__)


# =====================
# Equipment Profiles
# =====================

EQUIPMENT_PROFILES = {
    'bwms': {
        'name': 'BWMS (Ballast Water Management System)',
        'name_ko': '선박 평형수 처리 시스템',
        'description': '선박 평형수 처리 장비 태그 인식',
        'context_keywords': [
            'BWMS', 'BWTS', 'BALLAST', 'SIGNAL FOR BWMS',
            'HYCHLOR', 'ECS', 'ELECTROCHLOR'
        ],
        'equipment': {
            'ECU': {
                'pattern': r'ECU[-_\s]?\d*\w*',
                'name_ko': '전기분해 유닛',
                'name_en': 'Electrolyzer Cell Unit',
                'description': '해수를 전기분해하여 차아염소산나트륨(NaOCl) 생성',
                'category': 'core'
            },
            'HGU': {
                'pattern': r'HGU[-_\s]?\d*\w*',
                'name_ko': '수소가스 유닛',
                'name_en': 'Hydrogen Gas Unit',
                'description': '전기분해 과정에서 발생하는 수소가스 처리',
                'category': 'safety'
            },
            'FMU': {
                'pattern': r'FMU[-_\s]?\d*\w*',
                'name_ko': '필터 모듈',
                'name_en': 'Filter Module Unit',
                'description': '해수 내 부유물질 제거',
                'category': 'pretreatment'
            },
            'ANU': {
                'pattern': r'ANU[-_\s]?\d*\w*',
                'name_ko': '중화 유닛',
                'name_en': 'Active Neutralization Unit',
                'description': '처리수 중화 처리',
                'category': 'post_treatment'
            },
            'NIU': {
                'pattern': r'NIU[-_\s]?\d*\w*',
                'name_ko': '중화 주입 유닛',
                'name_en': 'Neutralization Injection Unit',
                'description': '중화제 주입',
                'category': 'post_treatment'
            },
            'TSU': {
                'pattern': r'TSU[-_\s]?\w*',
                'name_ko': '잔류 용액 유닛',
                'name_en': 'Total residual Solution Unit',
                'description': 'TRO(잔류산화제) 측정',
                'category': 'monitoring'
            },
            'DTS': {
                'pattern': r'DTS[-_\s]?\d*\w*',
                'name_ko': 'TRO 투여 스테이션',
                'name_en': 'Dosing TRO Station',
                'description': 'TRO 투여량 조절',
                'category': 'dosing'
            },
            'GDS': {
                'pattern': r'GDS[-_\s]?\d*\w*',
                'name_ko': '가스 희석 시스템',
                'name_en': 'Gas Dilution System',
                'description': '수소가스 안전 희석',
                'category': 'safety'
            },
            'EWU': {
                'pattern': r'EWU[-_\s]?\w*',
                'name_ko': '전해질 세척 유닛',
                'name_en': 'Electrolyte Washing Unit',
                'description': '전극 세척',
                'category': 'maintenance'
            },
            'APU': {
                'pattern': r'APU[-_\s]?\d*\w*',
                'name_ko': '자동 퍼지 유닛',
                'name_en': 'Automatic Purge Unit',
                'description': '시스템 자동 퍼지',
                'category': 'maintenance'
            },
            'DMU': {
                'pattern': r'DMU[-_\s]?\d*\w*',
                'name_ko': '직접 혼합 유닛',
                'name_en': 'Direct Mix Unit',
                'description': '처리수 직접 혼합',
                'category': 'mixing'
            },
            'CPC': {
                'pattern': r'CPC[-_\s]?\d*\w*',
                'name_ko': '제어 패널',
                'name_en': 'Control Panel Cabinet',
                'description': '시스템 제어 패널',
                'category': 'control'
            },
            'TRO': {
                'pattern': r'TRO[-_\s]?\w*',
                'name_ko': 'TRO 센서',
                'name_en': 'TRO Sensor',
                'description': '잔류산화제 농도 센서',
                'category': 'monitoring'
            },
            'PCU': {
                'pattern': r'PCU[-_\s]?\d*\w*',
                'name_ko': '전력 제어 유닛',
                'name_en': 'Power Control Unit',
                'description': '전원 공급 및 제어',
                'category': 'control'
            },
        }
    },
    'hvac': {
        'name': 'HVAC (Heating, Ventilation, and Air Conditioning)',
        'name_ko': '공조 시스템',
        'description': '건물/선박 공조 장비 태그 인식',
        'context_keywords': [
            'HVAC', 'AIR HANDLING', 'VENTILATION', 'COOLING', 'HEATING'
        ],
        'equipment': {
            'AHU': {
                'pattern': r'AHU[-_\s]?\d*\w*',
                'name_ko': '공조기',
                'name_en': 'Air Handling Unit',
                'description': '공기 조화 및 처리',
                'category': 'air_handling'
            },
            'FCU': {
                'pattern': r'FCU[-_\s]?\d*\w*',
                'name_ko': '팬 코일 유닛',
                'name_en': 'Fan Coil Unit',
                'description': '개별 공간 냉난방',
                'category': 'terminal'
            },
            'VAV': {
                'pattern': r'VAV[-_\s]?\d*\w*',
                'name_ko': '가변풍량 유닛',
                'name_en': 'Variable Air Volume',
                'description': '풍량 가변 조절',
                'category': 'terminal'
            },
            'EAF': {
                'pattern': r'EAF[-_\s]?\d*\w*',
                'name_ko': '배기팬',
                'name_en': 'Exhaust Air Fan',
                'description': '공기 배출',
                'category': 'fan'
            },
            'SAF': {
                'pattern': r'SAF[-_\s]?\d*\w*',
                'name_ko': '급기팬',
                'name_en': 'Supply Air Fan',
                'description': '공기 공급',
                'category': 'fan'
            },
        }
    },
    'process': {
        'name': 'General Process',
        'name_ko': '일반 공정',
        'description': '범용 공정 장비 태그 인식',
        'context_keywords': [],
        'equipment': {
            'P': {
                'pattern': r'P[-_]?\d{3,4}[A-Z]?',
                'name_ko': '펌프',
                'name_en': 'Pump',
                'description': '유체 이송',
                'category': 'rotating'
            },
            'V': {
                'pattern': r'V[-_]?\d{3,4}[A-Z]?',
                'name_ko': '용기/탱크',
                'name_en': 'Vessel/Tank',
                'description': '유체 저장',
                'category': 'vessel'
            },
            'E': {
                'pattern': r'E[-_]?\d{3,4}[A-Z]?',
                'name_ko': '열교환기',
                'name_en': 'Heat Exchanger',
                'description': '열 교환',
                'category': 'heat_transfer'
            },
            'C': {
                'pattern': r'C[-_]?\d{3,4}[A-Z]?',
                'name_ko': '압축기',
                'name_en': 'Compressor',
                'description': '가스 압축',
                'category': 'rotating'
            },
        }
    }
}


# =====================
# Core Functions
# =====================

def get_available_profiles() -> List[Dict]:
    """사용 가능한 장비 프로파일 목록 반환"""
    profiles = []
    for profile_id, profile_data in EQUIPMENT_PROFILES.items():
        profiles.append({
            'id': profile_id,
            'name': profile_data['name'],
            'name_ko': profile_data['name_ko'],
            'description': profile_data['description'],
            'equipment_count': len(profile_data['equipment'])
        })
    return profiles


def get_profile_equipment_types(profile_id: str) -> List[Dict]:
    """특정 프로파일의 장비 타입 목록 반환"""
    profile = EQUIPMENT_PROFILES.get(profile_id)
    if not profile:
        return []

    equipment_types = []
    for equip_type, info in profile['equipment'].items():
        equipment_types.append({
            'type': equip_type,
            'name_ko': info['name_ko'],
            'name_en': info['name_en'],
            'description': info['description'],
            'category': info['category']
        })
    return equipment_types


def detect_equipment(
    ocr_results: List[Dict],
    profile_id: str = 'bwms'
) -> List[Dict]:
    """
    OCR 결과에서 장비 태그를 검출합니다.

    Args:
        ocr_results: OCR 결과 리스트
            - 형식: [{'text': 'ECU-001', 'confidence': 0.95, 'bbox': [...], ...}, ...]
            - 또는 단순 텍스트 리스트: ['ECU-001', 'FMU-002', ...]
        profile_id: 사용할 장비 프로파일 (기본: 'bwms')

    Returns:
        검출된 장비 리스트
    """
    profile = EQUIPMENT_PROFILES.get(profile_id)
    if not profile:
        logger.warning(f"Unknown profile: {profile_id}, using 'bwms'")
        profile = EQUIPMENT_PROFILES['bwms']

    equipment_dict = profile['equipment']
    equipment_list = []
    seen_tags = set()

    for ocr_item in ocr_results:
        # OCR 결과 형식 처리
        if isinstance(ocr_item, dict):
            text = ocr_item.get('text', '').strip()
            confidence = ocr_item.get('confidence', 0.0)
            bbox = ocr_item.get('bbox', ocr_item.get('position', None))
        else:
            text = str(ocr_item).strip()
            confidence = 1.0
            bbox = None

        text_upper = text.upper()
        maker_supply = '*' in text

        for equip_type, info in equipment_dict.items():
            if re.search(info['pattern'], text_upper, re.IGNORECASE):
                match = re.search(info['pattern'], text_upper, re.IGNORECASE)
                if match:
                    tag = match.group(0).strip()
                    normalized_tag = re.sub(r'[\s_-]+', '-', tag).upper()

                    if normalized_tag not in seen_tags:
                        seen_tags.add(normalized_tag)
                        equipment_list.append({
                            'tag': tag,
                            'type': equip_type,
                            'name_ko': info['name_ko'],
                            'name_en': info['name_en'],
                            'description': info['description'],
                            'category': info['category'],
                            'bbox': bbox,
                            'confidence': confidence,
                            'maker_supply': maker_supply,
                            'original_text': text,
                            'profile': profile_id
                        })
                break

    equipment_list.sort(key=lambda x: (x['type'], x['tag']))
    logger.info(f"장비 검출 ({profile_id}): {len(equipment_list)}개")
    return equipment_list


def get_equipment_summary(equipment_list: List[Dict]) -> Dict:
    """장비 요약 통계 생성"""
    summary = {
        'total_count': len(equipment_list),
        'by_type': {},
        'by_category': {},
        'maker_supply_count': 0,
        'equipment_types': [],
        'profile': equipment_list[0]['profile'] if equipment_list else None
    }

    for equip in equipment_list:
        equip_type = equip['type']
        category = equip.get('category', 'unknown')

        if equip_type not in summary['by_type']:
            summary['by_type'][equip_type] = {
                'count': 0,
                'name_ko': equip['name_ko'],
                'name_en': equip['name_en'],
                'tags': []
            }
        summary['by_type'][equip_type]['count'] += 1
        summary['by_type'][equip_type]['tags'].append(equip['tag'])

        summary['by_category'][category] = summary['by_category'].get(category, 0) + 1

        if equip.get('maker_supply'):
            summary['maker_supply_count'] += 1

    summary['equipment_types'] = list(summary['by_type'].keys())
    return summary


def check_profile_context(
    ocr_results: List[Dict],
    profile_id: Optional[str] = None
) -> Dict:
    """
    도면이 특정 프로파일에 해당하는지 확인

    Args:
        ocr_results: OCR 결과
        profile_id: 확인할 프로파일 (None이면 모든 프로파일 확인)

    Returns:
        매칭된 프로파일 정보
    """
    all_texts = []
    for ocr_item in ocr_results:
        if isinstance(ocr_item, dict):
            text = ocr_item.get('text', '').upper()
        else:
            text = str(ocr_item).upper()
        all_texts.append(text)

    full_text = ' '.join(all_texts)

    if profile_id:
        profiles_to_check = {profile_id: EQUIPMENT_PROFILES.get(profile_id)}
    else:
        profiles_to_check = EQUIPMENT_PROFILES

    matches = []
    for pid, profile in profiles_to_check.items():
        if not profile:
            continue

        found_keywords = []
        for keyword in profile.get('context_keywords', []):
            if keyword in full_text:
                found_keywords.append(keyword)

        if found_keywords:
            confidence = min(len(found_keywords) / 3, 1.0)
            matches.append({
                'profile_id': pid,
                'name': profile['name'],
                'name_ko': profile['name_ko'],
                'confidence': round(confidence, 2),
                'indicators': found_keywords
            })

    matches.sort(key=lambda x: x['confidence'], reverse=True)

    return {
        'matched_profiles': matches,
        'best_match': matches[0] if matches else None,
        'has_match': len(matches) > 0
    }


def generate_equipment_list_excel(
    equipment_list: List[Dict],
    project_info: Optional[Dict] = None,
    profile_id: str = 'bwms'
) -> bytes:
    """
    Equipment List Excel 파일 생성

    Args:
        equipment_list: 장비 목록
        project_info: 프로젝트 정보
        profile_id: 프로파일 ID (Excel 시트 제목에 사용)

    Returns:
        Excel 파일 바이트
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from io import BytesIO
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        raise ImportError("openpyxl is required for Excel generation")

    if project_info is None:
        project_info = {}

    profile = EQUIPMENT_PROFILES.get(profile_id, EQUIPMENT_PROFILES['bwms'])
    profile_name = profile.get('name_ko', profile_id.upper())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment List"

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 헤더
    ws['A1'] = f"{profile_name} Equipment List"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')

    ws['A2'] = f"Project: {project_info.get('name', 'N/A')}"
    ws['A3'] = f"Drawing: {project_info.get('drawing_no', project_info.get('hull_no', 'N/A'))}"
    ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['D4'] = f"Total: {len(equipment_list)} items"

    # 테이블 헤더
    headers = ['No', 'Tag', 'Type', 'Name (Korean)', 'Name (English)', 'Category', 'Maker Supply']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # 데이터
    for i, equip in enumerate(equipment_list, 1):
        row = 6 + i
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=equip['tag']).border = border
        ws.cell(row=row, column=3, value=equip['type']).border = border
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=4, value=equip['name_ko']).border = border
        ws.cell(row=row, column=5, value=equip['name_en']).border = border
        ws.cell(row=row, column=6, value=equip.get('category', '')).border = border
        ws.cell(row=row, column=6).alignment = center_align
        ws.cell(row=row, column=7, value='*' if equip.get('maker_supply') else '').border = border
        ws.cell(row=row, column=7).alignment = center_align

    # 열 너비
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

    # 요약
    summary = get_equipment_summary(equipment_list)
    summary_row = 6 + len(equipment_list) + 2

    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    ws.merge_cells(f'A{summary_row}:B{summary_row}')

    ws.cell(row=summary_row + 1, column=1, value="Equipment by Type:")
    row_offset = 2
    for equip_type, type_info in summary['by_type'].items():
        ws.cell(row=summary_row + row_offset, column=2,
                value=f"{equip_type}: {type_info['count']} ({type_info['name_ko']})")
        row_offset += 1

    ws.cell(row=summary_row + row_offset, column=1,
            value=f"Maker Supply: {summary['maker_supply_count']}")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def extract_signal_region_equipment(
    symbols: List[Dict],
    regions: List[Dict],
    ocr_results: List[Dict],
    region_keywords: List[str] = None
) -> List[Dict]:
    """
    특정 영역 (점선 박스) 내의 장비/심볼 추출

    Args:
        symbols: YOLO 검출 심볼 (model_type=pid_class_aware)
        regions: Line Detector 검출 영역
        ocr_results: OCR 결과
        region_keywords: 영역 식별 키워드 (예: ['SIGNAL', 'BWMS'])

    Returns:
        영역 내 심볼 목록
    """
    if region_keywords is None:
        region_keywords = ['SIGNAL']

    signal_symbols = []
    matched_regions = []

    # 키워드가 포함된 영역 찾기
    for region in regions:
        region_bbox = region.get('bbox', [])
        if not region_bbox or len(region_bbox) < 4:
            continue

        for ocr_item in ocr_results:
            if isinstance(ocr_item, dict):
                text = ocr_item.get('text', '').upper()
            else:
                continue

            if all(kw.upper() in text for kw in region_keywords):
                matched_regions.append({
                    'bbox': region_bbox,
                    'type': region.get('type', 'signal_group'),
                    'label': text
                })
                break

    # 영역 내 심볼 추출
    for region in matched_regions:
        rx1, ry1, rx2, ry2 = region['bbox']

        for symbol in symbols:
            sym_center = symbol.get('center', [0, 0])
            sx, sy = sym_center

            if rx1 <= sx <= rx2 and ry1 <= sy <= ry2:
                signal_symbols.append({
                    'symbol_id': symbol.get('id'),
                    'class_name': symbol.get('class_name', 'unknown'),
                    'korean_name': symbol.get('korean_name', ''),
                    'center': sym_center,
                    'bbox': symbol.get('bbox', []),
                    'confidence': symbol.get('confidence', 0),
                    'region': region['label'],
                    'region_type': region['type']
                })

    logger.info(f"신호 영역 내 심볼 추출: {len(signal_symbols)}개")
    return signal_symbols


# =====================
# Utility Functions
# =====================

def format_equipment_for_display(equipment_list: List[Dict]) -> str:
    """장비 목록을 표시용 텍스트로 포맷"""
    if not equipment_list:
        return "검출된 장비 없음"

    profile_id = equipment_list[0].get('profile', 'unknown')
    profile = EQUIPMENT_PROFILES.get(profile_id, {})
    profile_name = profile.get('name_ko', profile_id)

    lines = [f"{profile_name} Equipment List", "=" * 50]

    for i, equip in enumerate(equipment_list, 1):
        maker = " *" if equip.get('maker_supply') else ""
        lines.append(
            f"{i:2}. [{equip['type']}] {equip['tag']}{maker}"
            f"\n     {equip['name_ko']} ({equip['name_en']})"
        )

    summary = get_equipment_summary(equipment_list)
    lines.append("")
    lines.append(f"Total: {summary['total_count']}")
    lines.append(f"Maker Supply: {summary['maker_supply_count']}")

    return "\n".join(lines)


# =====================
# Legacy Compatibility (BWMS)
# =====================

# 기존 BWMS 함수명 호환성 유지
def detect_bwms_equipment(ocr_results: List[Dict]) -> List[Dict]:
    """[Legacy] BWMS 장비 검출 - detect_equipment(profile_id='bwms') 사용 권장"""
    return detect_equipment(ocr_results, profile_id='bwms')


def get_bwms_equipment_summary(equipment_list: List[Dict]) -> Dict:
    """[Legacy] BWMS 요약 - get_equipment_summary() 사용 권장"""
    return get_equipment_summary(equipment_list)


def generate_bwms_equipment_list_excel(
    equipment_list: List[Dict],
    project_info: Optional[Dict] = None
) -> bytes:
    """[Legacy] BWMS Excel - generate_equipment_list_excel(profile_id='bwms') 사용 권장"""
    return generate_equipment_list_excel(equipment_list, project_info, profile_id='bwms')


def check_bwms_context(ocr_results: List[Dict]) -> Dict:
    """[Legacy] BWMS 컨텍스트 확인 - check_profile_context(profile_id='bwms') 사용 권장"""
    result = check_profile_context(ocr_results, profile_id='bwms')
    best = result.get('best_match', {})
    return {
        'is_bwms': result['has_match'],
        'confidence': best.get('confidence', 0),
        'indicators': best.get('indicators', [])
    }

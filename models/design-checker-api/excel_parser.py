"""
Excel Parser for BWMS Checklist Rules
TECHCROSS 체크리스트 Excel 파일을 파싱하여 규칙으로 변환

지원 형식:
- .xlsx (Excel 2007+)
- .xls (Excel 97-2003)
"""

import logging
import json
import yaml
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from enum import Enum
import re

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class CheckType(str, Enum):
    """검사 유형"""
    POSITION = "position"      # 위치 검증 (상류/하류, 상부/하부)
    SEQUENCE = "sequence"      # 순서 검증 (A 다음에 B)
    DISTANCE = "distance"      # 거리 검증 (≤5m)
    COUNT = "count"            # 개수 검증
    CAPACITY = "capacity"      # 용량 검증 (비율)
    EXISTENCE = "existence"    # 존재 여부
    CONNECTION = "connection"  # 연결 검증
    MANUAL = "manual"          # 수동 검토


class Severity(str, Enum):
    """심각도"""
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"


@dataclass
class ChecklistRule:
    """체크리스트 규칙"""
    rule_id: str
    name: str
    name_en: Optional[str] = None
    severity: str = "warning"
    check_type: str = "manual"
    category: str = "bwms"
    equipment: Optional[str] = None
    condition: Optional[str] = None
    condition_value: Optional[str] = None
    na_condition: Optional[str] = None
    description: Optional[str] = None
    suggestion: Optional[str] = None
    auto_checkable: bool = False
    standard: str = "TECHCROSS BWMS Standard"
    product_type: Optional[str] = None  # ECS, HYCHLOR, ALL

    def to_dict(self) -> Dict[str, Any]:
        return {k: v for k, v in asdict(self).items() if v is not None}


class ExcelRuleParser:
    """Excel 체크리스트 규칙 파서"""

    # 필수 컬럼
    REQUIRED_COLUMNS = ["rule_id", "name", "severity", "check_type"]

    # 컬럼 매핑 (한글 → 영문)
    COLUMN_MAPPING = {
        # 한글
        "규칙ID": "rule_id",
        "규칙 ID": "rule_id",
        "ID": "rule_id",
        "번호": "rule_id",
        "No": "rule_id",
        "No.": "rule_id",

        "규칙명": "name",
        "규칙 이름": "name",
        "항목": "name",
        "체크항목": "name",
        "검사항목": "name",
        "Name": "name",

        "영문명": "name_en",
        "English": "name_en",

        "심각도": "severity",
        "중요도": "severity",
        "Severity": "severity",
        "Level": "severity",

        "검사유형": "check_type",
        "검사 유형": "check_type",
        "유형": "check_type",
        "Type": "check_type",
        "Check Type": "check_type",

        "장비": "equipment",
        "대상장비": "equipment",
        "Equipment": "equipment",

        "조건": "condition",
        "검사조건": "condition",
        "Condition": "condition",

        "조건값": "condition_value",
        "값": "condition_value",
        "Value": "condition_value",

        "N/A조건": "na_condition",
        "NA조건": "na_condition",
        "N/A Condition": "na_condition",
        "해당없음조건": "na_condition",

        "설명": "description",
        "Description": "description",
        "비고": "description",

        "권장조치": "suggestion",
        "조치": "suggestion",
        "Suggestion": "suggestion",

        "자동화": "auto_checkable",
        "자동검사": "auto_checkable",
        "Auto": "auto_checkable",
        "Automatable": "auto_checkable",

        "표준": "standard",
        "Standard": "standard",

        "제품": "product_type",
        "제품유형": "product_type",
        "제품타입": "product_type",
        "Product": "product_type",
        "Product Type": "product_type",
    }

    # 심각도 매핑
    SEVERITY_MAPPING = {
        "error": "error", "에러": "error", "오류": "error", "높음": "error", "high": "error", "필수": "error",
        "warning": "warning", "경고": "warning", "주의": "warning", "중간": "warning", "medium": "warning",
        "info": "info", "정보": "info", "참고": "info", "낮음": "info", "low": "info",
    }

    # 검사유형 매핑
    CHECK_TYPE_MAPPING = {
        "position": "position", "위치": "position", "배치": "position",
        "sequence": "sequence", "순서": "sequence", "흐름": "sequence",
        "distance": "distance", "거리": "distance", "이격": "distance",
        "count": "count", "개수": "count", "수량": "count",
        "capacity": "capacity", "용량": "capacity", "비율": "capacity",
        "existence": "existence", "존재": "existence", "유무": "existence",
        "connection": "connection", "연결": "connection", "접속": "connection",
        "manual": "manual", "수동": "manual", "검토": "manual",
    }

    def __init__(self, config_dir: str = "config"):
        self.config_dir = Path(config_dir)
        self.config_dir.mkdir(parents=True, exist_ok=True)

    def parse_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[ChecklistRule], List[str]]:
        """
        Excel 파일을 파싱하여 규칙 목록 반환

        Returns:
            Tuple[List[ChecklistRule], List[str]]: (규칙 목록, 경고 메시지)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl이 설치되지 않았습니다. pip install openpyxl")

        rules = []
        warnings = []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # 시트 선택
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다. 사용 가능: {wb.sheetnames}")
                ws = wb[sheet_name]
            else:
                # 첫 번째 시트 또는 'Rules', '규칙' 등의 이름을 가진 시트 찾기
                ws = None
                for name in ["Rules", "규칙", "체크리스트", "Checklist"]:
                    if name in wb.sheetnames:
                        ws = wb[name]
                        break
                if ws is None:
                    ws = wb.active

            # 헤더 행 찾기 (첫 번째 비어있지 않은 행)
            header_row = None
            headers = {}

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
                cells = [cell.value for cell in row]
                if any(cells):
                    # 헤더 후보인지 확인 (rule_id 또는 번호 등이 있는지)
                    for col_idx, cell_value in enumerate(cells):
                        if cell_value:
                            normalized = self._normalize_column_name(str(cell_value))
                            if normalized in self.REQUIRED_COLUMNS:
                                header_row = row_idx
                                break

                    if header_row:
                        # 헤더 매핑
                        for col_idx, cell_value in enumerate(cells):
                            if cell_value:
                                normalized = self._normalize_column_name(str(cell_value))
                                headers[col_idx] = normalized
                        break

            if not header_row:
                raise ValueError("헤더 행을 찾을 수 없습니다. 필수 컬럼: " + ", ".join(self.REQUIRED_COLUMNS))

            # 필수 컬럼 확인
            found_columns = set(headers.values())
            missing = set(self.REQUIRED_COLUMNS) - found_columns
            if missing:
                warnings.append(f"필수 컬럼 누락: {missing}. 기본값이 사용됩니다.")

            # 데이터 행 파싱
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
                cells = [cell.value for cell in row]

                # 빈 행 건너뛰기
                if not any(cells):
                    continue

                # 규칙 ID가 없으면 건너뛰기
                rule_id = None
                for col_idx, col_name in headers.items():
                    if col_name == "rule_id" and col_idx < len(cells):
                        rule_id = cells[col_idx]
                        break

                if not rule_id:
                    continue

                # 규칙 생성
                rule_data = {"rule_id": str(rule_id)}

                for col_idx, col_name in headers.items():
                    if col_idx < len(cells) and cells[col_idx] is not None:
                        value = cells[col_idx]

                        # 값 변환
                        if col_name == "severity":
                            value = self._normalize_severity(str(value))
                        elif col_name == "check_type":
                            value = self._normalize_check_type(str(value))
                        elif col_name == "auto_checkable":
                            value = self._normalize_bool(value)
                        else:
                            value = str(value).strip() if value else None

                        if value:
                            rule_data[col_name] = value

                try:
                    rule = ChecklistRule(**rule_data)
                    rules.append(rule)
                except Exception as e:
                    warnings.append(f"행 {row_idx}: 규칙 생성 실패 - {e}")

            wb.close()

        except Exception as e:
            raise ValueError(f"Excel 파일 파싱 실패: {e}")

        logger.info(f"Excel 파싱 완료: {len(rules)}개 규칙, {len(warnings)}개 경고")
        return rules, warnings

    def _normalize_column_name(self, name: str) -> str:
        """컬럼명 정규화"""
        name = name.strip()

        # 줄바꿈이 있으면 첫 줄만 사용 (템플릿 형식: "규칙ID\n(Rule ID)")
        if "\n" in name:
            name = name.split("\n")[0].strip()

        # 괄호 제거 (영문만 있는 경우)
        if name.startswith("(") and name.endswith(")"):
            name = name[1:-1].strip()

        return self.COLUMN_MAPPING.get(name, name.lower().replace(" ", "_"))

    def _normalize_severity(self, value: str) -> str:
        """심각도 정규화"""
        value = value.strip().lower()
        return self.SEVERITY_MAPPING.get(value, "warning")

    def _normalize_check_type(self, value: str) -> str:
        """검사유형 정규화"""
        value = value.strip().lower()
        return self.CHECK_TYPE_MAPPING.get(value, "manual")

    def _normalize_bool(self, value: Any) -> bool:
        """불리언 정규화"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in ("true", "yes", "y", "1", "o", "○", "●", "✓", "✔")
        return bool(value)

    def save_rules_yaml(self, rules: List[ChecklistRule], filename: str) -> str:
        """규칙을 YAML 파일로 저장"""
        filepath = self.config_dir / filename

        data = {
            "version": "1.0",
            "product_type": rules[0].product_type if rules else "ALL",
            "rules_count": len(rules),
            "rules": [rule.to_dict() for rule in rules]
        }

        with open(filepath, "w", encoding="utf-8") as f:
            yaml.dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

        logger.info(f"YAML 저장 완료: {filepath}")
        return str(filepath)

    def save_rules_json(self, rules: List[ChecklistRule], filename: str) -> str:
        """규칙을 JSON 파일로 저장"""
        filepath = self.config_dir / filename

        data = {
            "version": "1.0",
            "product_type": rules[0].product_type if rules else "ALL",
            "rules_count": len(rules),
            "rules": [rule.to_dict() for rule in rules]
        }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON 저장 완료: {filepath}")
        return str(filepath)

    def load_rules_yaml(self, filename: str) -> List[ChecklistRule]:
        """YAML 파일에서 규칙 로드"""
        filepath = self.config_dir / filename

        if not filepath.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

        with open(filepath, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)

        rules = []
        for rule_data in data.get("rules", []):
            rules.append(ChecklistRule(**rule_data))

        return rules

    def load_rules_json(self, filename: str) -> List[ChecklistRule]:
        """JSON 파일에서 규칙 로드"""
        filepath = self.config_dir / filename

        if not filepath.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        rules = []
        for rule_data in data.get("rules", []):
            rules.append(ChecklistRule(**rule_data))

        return rules

    def list_rule_files(self) -> List[Dict[str, Any]]:
        """저장된 규칙 파일 목록"""
        files = []

        for ext in ["*.yaml", "*.yml", "*.json"]:
            for filepath in self.config_dir.glob(ext):
                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        if filepath.suffix in [".yaml", ".yml"]:
                            data = yaml.safe_load(f)
                        else:
                            data = json.load(f)

                    files.append({
                        "filename": filepath.name,
                        "path": str(filepath),
                        "version": data.get("version", "unknown"),
                        "product_type": data.get("product_type", "unknown"),
                        "rules_count": data.get("rules_count", len(data.get("rules", []))),
                        "modified": filepath.stat().st_mtime
                    })
                except Exception as e:
                    logger.warning(f"파일 로드 실패: {filepath} - {e}")

        return sorted(files, key=lambda x: x["modified"], reverse=True)


def create_excel_template(output_path: str = "templates/bwms_checklist_template.xlsx") -> str:
    """
    BWMS 체크리스트 Excel 템플릿 생성
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl이 설치되지 않았습니다. pip install openpyxl")

    # 디렉토리 생성
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더 정의
    headers = [
        ("Rule ID", "규칙ID", 12),
        ("Name", "규칙명", 40),
        ("Name (EN)", "영문명", 30),
        ("Severity", "심각도", 12),
        ("Check Type", "검사유형", 15),
        ("Equipment", "대상장비", 20),
        ("Condition", "검사조건", 30),
        ("Value", "조건값", 15),
        ("N/A Condition", "N/A조건", 25),
        ("Description", "설명", 40),
        ("Suggestion", "권장조치", 30),
        ("Auto", "자동화", 10),
        ("Product", "제품", 12),
    ]

    # 헤더 행 작성
    for col_idx, (en_name, kr_name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=f"{kr_name}\n({en_name})")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 예시 데이터
    sample_rules = [
        ["BWMS-001", "G-2 Sampling Port 상류 위치", "G-2 Sampling Port Upstream Position",
         "warning", "position", "G-2 Sampling Port", "upstream_of", "Main Line",
         "", "G-2 Sampling Port는 주 배관 상류에 위치해야 합니다",
         "Sampling Port를 상류로 이동하세요", "O", "ALL"],
        ["BWMS-004", "FMU-ECU 순서 검증", "FMU-ECU Sequence Validation",
         "error", "sequence", "FMU,ECU", "after", "ECU",
         "", "FMU(유량계)는 ECU(전해조) 후단에 위치해야 합니다",
         "FMU를 ECU 후단(downstream)에 배치하세요", "O", "ALL"],
        ["BWMS-005", "GDS 위치 검증", "GDS Position Validation",
         "error", "position", "GDS", "above", "ECU,HGU",
         "", "GDS(가스감지센서)는 ECU 또는 HGU 상부에 위치해야 합니다",
         "GDS를 ECU 또는 HGU 상부에 배치하세요", "O", "ALL"],
        ["BWMS-006", "TSU-APU 거리 제한", "TSU-APU Distance Limit",
         "warning", "distance", "TSU,APU", "max_distance", "5m",
         "No Scale Info", "TSU와 APU 간 거리는 5m 이내여야 합니다",
         "TSU와 APU 거리를 5m 이내로 조정하세요", "△", "ALL"],
        ["BWMS-007", "Mixing Pump 용량 검증", "Mixing Pump Capacity Check",
         "warning", "capacity", "Mixing Pump", "ratio", "4.3%",
         "", "Mixing Pump 용량 = Ballast Pump × 4.3%",
         "Mixing Pump 용량을 확인하세요", "△", "ECS"],
        ["BWMS-009", "HYCHLOR 필터 위치", "HYCHLOR Filter Position",
         "warning", "position", "Filter", "before", "HGU",
         "ECS 모델", "HYCHLOR 필터는 HGU 전단에 위치해야 합니다",
         "필터를 HGU 전단에 배치하세요", "O", "HYCHLOR"],
    ]

    # 데이터 행 작성
    for row_idx, rule in enumerate(sample_rules, start=2):
        for col_idx, value in enumerate(rule, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 행 높이 설정
    ws.row_dimensions[1].height = 40
    for row_idx in range(2, len(sample_rules) + 2):
        ws.row_dimensions[row_idx].height = 30

    # 설명 시트 추가
    ws_info = wb.create_sheet("설명")
    info_data = [
        ["컬럼", "설명", "예시값"],
        ["Rule ID", "규칙 고유 식별자", "BWMS-001, BWMS-004"],
        ["Name", "규칙명 (한글)", "FMU-ECU 순서 검증"],
        ["Name (EN)", "규칙명 (영문)", "FMU-ECU Sequence Validation"],
        ["Severity", "심각도: error, warning, info", "error"],
        ["Check Type", "검사유형: position, sequence, distance, count, capacity, existence, connection, manual", "sequence"],
        ["Equipment", "대상 장비 (쉼표로 구분)", "FMU,ECU"],
        ["Condition", "검사 조건: upstream_of, downstream_of, above, below, before, after, max_distance, ratio", "after"],
        ["Value", "조건값", "ECU, 5m, 4.3%"],
        ["N/A Condition", "해당없음 조건", "ECS 모델, No Scale Info"],
        ["Description", "규칙 설명", "FMU는 ECU 후단에 위치해야 합니다"],
        ["Suggestion", "위반 시 권장 조치", "FMU를 ECU 후단에 배치하세요"],
        ["Auto", "자동 검사 가능 여부: O(가능), △(부분), X(불가)", "O"],
        ["Product", "적용 제품: ECS, HYCHLOR, ALL", "ALL"],
    ]

    for row_idx, row_data in enumerate(info_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws_info.column_dimensions['A'].width = 15
    ws_info.column_dimensions['B'].width = 60
    ws_info.column_dimensions['C'].width = 40

    # 저장
    wb.save(output_path)
    logger.info(f"템플릿 생성 완료: {output_path}")
    return output_path


# 싱글톤 파서 인스턴스
excel_parser = ExcelRuleParser()

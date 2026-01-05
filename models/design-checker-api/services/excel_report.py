"""
Excel Report Service - TECHCROSS 체크리스트 형식 Excel 리포트 생성
"""
import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 스타일 정의
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelReportService:
    """TECHCROSS 체크리스트 Excel 리포트 생성"""

    def generate_validation_report(
        self,
        validation_data: dict[str, Any],
        project_info: dict[str, Any] = None,
    ) -> bytes:
        """
        검증 결과를 Excel 리포트로 생성

        Args:
            validation_data: validate-with-mapping 엔드포인트 응답 데이터
            project_info: 프로젝트 정보 (선명, 도면번호 등)

        Returns:
            Excel 파일 바이트
        """
        wb = Workbook()

        # 1. 요약 시트
        self._create_summary_sheet(wb, validation_data, project_info)

        # 2. 통과 항목 시트
        self._create_passed_sheet(wb, validation_data.get("passed", []))

        # 3. 위반 항목 시트
        self._create_violations_sheet(wb, validation_data.get("violations", []))

        # 4. 스킵 항목 시트
        self._create_skipped_sheet(wb, validation_data.get("skipped", []))

        # 기본 시트 삭제 (있는 경우)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 바이트로 변환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_summary_sheet(
        self,
        wb: Workbook,
        validation_data: dict,
        project_info: dict = None
    ):
        """요약 시트 생성"""
        ws = wb.create_sheet("요약", 0)

        # 제목
        ws.merge_cells('A1:F1')
        ws['A1'] = "TECHCROSS BWMS P&ID 검증 리포트"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        # 생성 시간
        ws.merge_cells('A2:F2')
        ws['A2'] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # 프로젝트 정보
        row = 4
        if project_info:
            ws[f'A{row}'] = "프로젝트 정보"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            info_items = [
                ("선명", project_info.get("ship_name", "-")),
                ("도면번호", project_info.get("drawing_no", "-")),
                ("제품유형", project_info.get("product_type", "-")),
                ("선종", project_info.get("ship_type", "-")),
                ("선급", project_info.get("class_society", "-")),
            ]

            for label, value in info_items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

            row += 1

        # 필터 정보
        filters = validation_data.get("filters", {})
        ws[f'A{row}'] = "적용된 필터"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        filter_items = [
            ("제품유형", filters.get("product_type", "-")),
            ("선종", filters.get("ship_type", "-") or "-"),
            ("선급", filters.get("class_society", "-") or "-"),
            ("프로젝트", filters.get("project_type", "-") or "-"),
        ]

        for label, value in filter_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # 검증 결과 요약
        summary = validation_data.get("summary", {})
        ws[f'A{row}'] = "검증 결과 요약"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # 요약 테이블 헤더
        headers = ["항목", "수량", "비율"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        row += 1

        total = summary.get("existence_checked", 0)
        passed = summary.get("passed", 0)
        violations = summary.get("violations", 0)
        skipped = summary.get("skipped", 0)

        # 요약 데이터
        summary_rows = [
            ("총 규칙", summary.get("total_rules", 0), "-"),
            ("검사됨", total, "-"),
            ("통과", passed, f"{passed/total*100:.1f}%" if total > 0 else "-"),
            ("위반", violations, f"{violations/total*100:.1f}%" if total > 0 else "-"),
            ("스킵", skipped, "-"),
        ]

        for label, count, ratio in summary_rows:
            ws.cell(row=row, column=1, value=label).border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = THIN_BORDER
            ws.cell(row=row, column=3, value=ratio).border = THIN_BORDER

            # 색상
            if label == "통과":
                ws.cell(row=row, column=1).fill = PASS_FILL
            elif label == "위반":
                ws.cell(row=row, column=1).fill = FAIL_FILL
            elif label == "스킵":
                ws.cell(row=row, column=1).fill = SKIP_FILL

            row += 1

        # 통과율 강조
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        pass_rate = summary.get("pass_rate", 0)
        ws[f'A{row}'] = f"통과율: {pass_rate:.1f}%"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        if pass_rate >= 90:
            ws[f'A{row}'].fill = PASS_FILL
        elif pass_rate >= 70:
            ws[f'A{row}'].fill = SKIP_FILL
        else:
            ws[f'A{row}'].fill = FAIL_FILL

        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _create_passed_sheet(self, wb: Workbook, passed: list):
        """통과 항목 시트"""
        ws = wb.create_sheet("통과 항목")

        # 헤더
        headers = ["No", "규칙 ID", "규칙명", "장비", "매칭 방식", "검출 태그", "비고"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

        # 데이터
        for row_idx, item in enumerate(passed, 2):
            matched_tags = item.get("matched_tags", [])
            tags_str = ", ".join(matched_tags[:5]) if matched_tags else "-"
            if len(matched_tags) > 5:
                tags_str += f" 외 {len(matched_tags)-5}개"

            reason = item.get("reason", "")

            data = [
                row_idx - 1,
                item.get("rule_id", ""),
                item.get("rule_name", ""),
                item.get("equipment", ""),
                item.get("match_type", ""),
                tags_str,
                reason if reason else "-",
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = THIN_BORDER
                cell.fill = PASS_FILL

        # 열 너비 조정
        widths = [5, 15, 30, 20, 15, 40, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_violations_sheet(self, wb: Workbook, violations: list):
        """위반 항목 시트"""
        ws = wb.create_sheet("위반 항목")

        # 헤더
        headers = ["No", "규칙 ID", "규칙명", "장비", "심각도", "권장 조치", "비고"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

        # 데이터
        for row_idx, item in enumerate(violations, 2):
            severity = item.get("severity", "warning")
            severity_kr = {"error": "오류", "warning": "경고", "info": "정보"}.get(severity, severity)

            data = [
                row_idx - 1,
                item.get("rule_id", ""),
                item.get("rule_name", ""),
                item.get("equipment", ""),
                severity_kr,
                item.get("suggestion", ""),
                item.get("reason", "-"),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = THIN_BORDER
                cell.fill = FAIL_FILL

        # 열 너비 조정
        widths = [5, 15, 30, 20, 10, 40, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 위반 없으면 메시지
        if not violations:
            ws.merge_cells('A2:G2')
            ws['A2'] = "위반 항목이 없습니다."
            ws['A2'].fill = PASS_FILL
            ws['A2'].alignment = Alignment(horizontal='center')

    def _create_skipped_sheet(self, wb: Workbook, skipped: list):
        """스킵 항목 시트"""
        ws = wb.create_sheet("스킵 항목")

        # 헤더
        headers = ["No", "규칙 ID", "검사 유형", "카테고리", "스킵 사유"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

        # 데이터
        for row_idx, item in enumerate(skipped, 2):
            data = [
                row_idx - 1,
                item.get("rule_id", ""),
                item.get("check_type", ""),
                item.get("category", ""),
                item.get("reason", ""),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = THIN_BORDER
                cell.fill = SKIP_FILL

        # 열 너비 조정
        widths = [5, 15, 15, 15, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def generate_checklist_report(
        self,
        validation_data: dict[str, Any],
        all_rules: list[dict],
        project_info: dict[str, Any] = None,
    ) -> bytes:
        """
        TECHCROSS 체크리스트 형식 Excel 생성

        모든 규칙을 포함하고, 검사 결과/상태를 표시
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "BWMS 체크리스트"

        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = "TECHCROSS BWMS P&ID 체크리스트"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        # 프로젝트 정보
        row = 3
        if project_info:
            info_text = f"선명: {project_info.get('ship_name', '-')} | "
            info_text += f"도면: {project_info.get('drawing_no', '-')} | "
            info_text += f"제품: {project_info.get('product_type', '-')}"
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = info_text
            row += 1

        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 2

        # 헤더
        headers = ["No", "규칙 ID", "검사 항목", "장비", "상태", "결과", "검출 태그", "비고"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # 결과 맵 생성
        passed_map = {p.get("rule_id"): p for p in validation_data.get("passed", [])}
        violation_map = {v.get("rule_id"): v for v in validation_data.get("violations", [])}
        skipped_map = {s.get("rule_id"): s for s in validation_data.get("skipped", [])}

        # 모든 규칙 순회
        no = 1
        for rule in all_rules:
            rule_id = rule.get("rule_id", "")

            # 상태 결정
            if rule_id in passed_map:
                status = "검사됨"
                result = "통과"
                fill = PASS_FILL
                item = passed_map[rule_id]
                tags = ", ".join(item.get("matched_tags", [])[:3])
                note = item.get("reason", "")
            elif rule_id in violation_map:
                status = "검사됨"
                result = "위반"
                fill = FAIL_FILL
                item = violation_map[rule_id]
                tags = "-"
                note = item.get("suggestion", "")
            elif rule_id in skipped_map:
                status = "스킵"
                result = "-"
                fill = SKIP_FILL
                item = skipped_map[rule_id]
                tags = "-"
                note = item.get("reason", "")
            else:
                status = "미검사"
                result = "-"
                fill = None
                tags = "-"
                note = "자동 검증 미지원"

            data = [
                no,
                rule_id,
                rule.get("name", ""),
                rule.get("equipment", ""),
                status,
                result,
                tags,
                note,
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = fill

            row += 1
            no += 1

        # 열 너비 조정
        widths = [5, 15, 35, 20, 10, 8, 30, 30]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 바이트로 변환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()


# 싱글톤 인스턴스
excel_report_service = ExcelReportService()

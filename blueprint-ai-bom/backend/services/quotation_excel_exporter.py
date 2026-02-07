"""Quotation Excel Exporter - 견적서 Excel 생성

openpyxl 기반 Excel 견적서 내보내기 (프로젝트 전체 / 어셈블리 단위)
QuotationService에서 분리된 독립 모듈
"""

import logging
from pathlib import Path
from typing import Optional

from schemas.quotation import ProjectQuotationResponse

logger = logging.getLogger(__name__)


def export_excel(
    quotation_data: ProjectQuotationResponse,
    output_dir: Path,
    customer_name: Optional[str] = None,
    include_material_breakdown: bool = True,
    notes: Optional[str] = None,
) -> Path:
    """openpyxl 기반 견적서 Excel

    Args:
        quotation_data: 프로젝트 견적 집계 데이터
        output_dir: 출력 디렉토리
        customer_name: 고객명 (없으면 quotation_data.customer 사용)
        include_material_breakdown: 재질별 분류 포함 여부
        notes: 비고

    Returns:
        생성된 Excel 파일 경로
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    output_path = output_dir / f"quotation_{quotation_data.project_id}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = "견적서 (Quotation)"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 메타 정보
    effective_customer = customer_name or quotation_data.customer
    ws['A3'] = f"프로젝트: {quotation_data.project_name}"
    ws['A4'] = f"고객: {effective_customer}"
    ws['A5'] = f"생성일: {quotation_data.created_at[:10]}"

    # 헤더
    headers = ["No.", "도면번호", "설명", "재질", "수량", "소계"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for row_idx, item in enumerate(quotation_data.items, 8):
        ws.cell(row=row_idx, column=1, value=row_idx - 7).border = border
        ws.cell(row=row_idx, column=2,
                value=item.drawing_number or "-").border = border
        ws.cell(row=row_idx, column=3,
                value=item.description or "-").border = border
        ws.cell(row=row_idx, column=4,
                value=item.material or "-").border = border
        ws.cell(row=row_idx, column=5,
                value=item.bom_quantity).border = border
        ws.cell(row=row_idx, column=6,
                value=f"₩{item.subtotal:,.0f}" if item.subtotal else "-"
                ).border = border

    # 요약
    summary_row = 8 + len(quotation_data.items) + 1
    summary = quotation_data.summary
    bold_font = Font(bold=True, size=12)

    ws.cell(row=summary_row, column=5, value="소계").font = bold_font
    ws.cell(row=summary_row, column=6, value=f"₩{summary.subtotal:,.0f}")
    ws.cell(row=summary_row + 1, column=5, value="부가세 (10%)").font = bold_font
    ws.cell(row=summary_row + 1, column=6, value=f"₩{summary.vat:,.0f}")
    ws.cell(row=summary_row + 2, column=5, value="합계").font = Font(bold=True, size=14)
    ws.cell(row=summary_row + 2, column=6, value=f"₩{summary.total:,.0f}"
            ).font = Font(bold=True, size=14)

    # 재질별 분류 시트
    if include_material_breakdown and quotation_data.material_groups:
        ws2 = wb.create_sheet(title="재질별 분류")
        mat_headers = ["재질", "품목수", "총 수량", "소계"]
        for col, header in enumerate(mat_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, mg in enumerate(quotation_data.material_groups, 2):
            ws2.cell(row=row_idx, column=1, value=mg.material).border = border
            ws2.cell(row=row_idx, column=2, value=mg.item_count).border = border
            ws2.cell(row=row_idx, column=3, value=mg.total_quantity).border = border
            ws2.cell(row=row_idx, column=4,
                     value=f"₩{mg.subtotal:,.0f}" if mg.subtotal else "-"
                     ).border = border

    # 열 너비
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 14

    wb.save(output_path)
    return output_path


def export_assembly_excel(
    quotation_data: ProjectQuotationResponse,
    assembly_drawing_number: str,
    output_dir: Path,
    customer_name: Optional[str] = None,
    notes: Optional[str] = None,
) -> Path:
    """특정 어셈블리만 포함한 Excel 견적서

    Args:
        quotation_data: 프로젝트 견적 집계 데이터
        assembly_drawing_number: 대상 어셈블리 도면번호
        output_dir: 출력 디렉토리
        customer_name: 고객명 (없으면 quotation_data.customer 사용)
        notes: 비고

    Returns:
        생성된 Excel 파일 경로
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # 어셈블리 그룹 찾기
    assy_group = next(
        (g for g in quotation_data.assembly_groups
         if g.assembly_drawing_number == assembly_drawing_number),
        None
    )
    if not assy_group:
        raise ValueError(f"어셈블리를 찾을 수 없습니다: {assembly_drawing_number}")

    safe_assy = assembly_drawing_number.replace("/", "_").replace(" ", "_")
    output_path = output_dir / f"quotation_{quotation_data.project_id}_{safe_assy}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    header_fill = PatternFill(
        start_color="E91E8C", end_color="E91E8C", fill_type="solid"
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = f"견적서 - {assy_group.assembly_description or assembly_drawing_number}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 메타 정보
    effective_customer = customer_name or quotation_data.customer
    ws['A3'] = f"프로젝트: {quotation_data.project_name}"
    ws['A4'] = f"고객: {effective_customer}"
    ws['A5'] = f"어셈블리: {assembly_drawing_number}"
    ws['D3'] = f"생성일: {quotation_data.created_at[:10]}"
    ws['D4'] = f"부품수: {assy_group.total_parts}개"

    # 헤더
    headers = ["No.", "도면번호", "품명", "재질", "수량", "소계"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for row_idx, item in enumerate(assy_group.items, 8):
        ws.cell(row=row_idx, column=1, value=row_idx - 7).border = border
        ws.cell(row=row_idx, column=2, value=item.drawing_number or "-").border = border
        ws.cell(row=row_idx, column=3, value=item.description or "-").border = border
        ws.cell(row=row_idx, column=4, value=item.material or "-").border = border
        ws.cell(row=row_idx, column=5, value=item.bom_quantity).border = border
        ws.cell(row=row_idx, column=6,
                value=f"₩{item.subtotal:,.0f}" if item.subtotal else "-"
                ).border = border

    # 요약
    summary_row = 8 + len(assy_group.items) + 1
    bold_font = Font(bold=True, size=11)

    ws.cell(row=summary_row, column=5, value="소계").font = bold_font
    ws.cell(row=summary_row, column=6, value=f"₩{assy_group.subtotal:,.0f}")
    ws.cell(row=summary_row + 1, column=5, value="부가세 (10%)").font = bold_font
    ws.cell(row=summary_row + 1, column=6, value=f"₩{assy_group.vat:,.0f}")
    ws.cell(row=summary_row + 2, column=5, value="합계").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 2, column=6, value=f"₩{assy_group.total:,.0f}"
            ).font = Font(bold=True, size=12)

    # 열 너비
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 14

    wb.save(output_path)
    logger.info(f"어셈블리 Excel 생성: {output_path}")
    return output_path

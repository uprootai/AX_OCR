"""BOM Service - BOM 생성 및 내보내기 서비스"""

import json
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any
from collections import defaultdict

from schemas.bom import BOMItem, BOMData, BOMSummary, ExportFormat
from schemas.typed_dicts import PricingInfo, DetectionDict, DimensionDict, RelationDict
from services.utils.pricing_utils import load_pricing_db, get_pricing_info as _get_pricing_info

logger = logging.getLogger(__name__)


class BOMService:
    """BOM 생성 및 내보내기 서비스"""

    def __init__(self, output_dir: Path, pricing_db_path: Optional[str] = None):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.pricing_db = load_pricing_db(pricing_db_path or "/app/classes_info_with_pricing.json")

    def get_pricing_info(self, class_name: str) -> PricingInfo:
        """클래스별 가격 정보 조회"""
        # BOM 서비스는 기본값이 약간 다름 (단가 10000, 리드타임 14)
        info = dict(_get_pricing_info(self.pricing_db, class_name))
        if info.get("단가") == 0:
            info["단가"] = 10000
        if info.get("리드타임") == 0:
            info["리드타임"] = 14
        return info  # type: ignore[return-value]

    def generate_bom(
        self,
        session_id: str,
        detections: List[DetectionDict],
        dimensions: Optional[List[DimensionDict]] = None,
        links: Optional[List[RelationDict]] = None,
        filename: Optional[str] = None,
        model_id: Optional[str] = None,
        session_pricing_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """검출 결과로부터 BOM 생성"""

        # 세션별 단가 DB 로드 (있으면 세션 단가 우선, 없으면 글로벌 폴백)
        pricing_db = self.pricing_db
        if session_pricing_path and os.path.exists(session_pricing_path):
            session_pricing = load_pricing_db(session_pricing_path)
            if session_pricing:
                pricing_db = session_pricing
                logger.info(f"세션별 단가 사용: {session_pricing_path} ({len(session_pricing)}개)")
            else:
                logger.warning(f"세션 단가 파일 로드 실패, 글로벌 폴백: {session_pricing_path}")

        # 승인된 검출만 필터링
        approved_detections = [
            d for d in detections
            if d.get("verification_status") in ("approved", "modified", "manual")
        ]

        # 치수 및 링크 맵핑 준비
        dim_map = {d["id"]: d for d in (dimensions or [])}
        symbol_to_dims = defaultdict(list)
        
        if links and dimensions:
            for link in links:
                symbol_id = link.get("symbol_id")
                dim_id = link.get("dimension_id")
                
                # 링크가 유효하고, 심볼이 존재하며, 치수가 존재하는 경우
                if symbol_id and dim_id and dim_id in dim_map:
                    dim_data = dim_map[dim_id]
                    # 승인된 치수만 포함 (선택사항 - 현재는 모두 포함하되 상태 표시)
                    symbol_to_dims[symbol_id].append(dim_data)

        # 클래스별 그룹화
        grouped = defaultdict(list)
        for detection in approved_detections:
            # 수정된 클래스명이 있으면 사용
            class_name = detection.get("modified_class_name") or detection.get("class_name")
            grouped[class_name].append(detection)

        # BOM 항목 생성
        items = []
        item_no = 1

        for class_name, detections_list in sorted(grouped.items()):
            quantity = len(detections_list)
            avg_confidence = sum(d["confidence"] for d in detections_list) / quantity

            # 가격 정보 조회 (세션 단가 또는 글로벌 단가)
            raw_info = dict(_get_pricing_info(pricing_db, class_name))
            if raw_info.get("단가") == 0:
                raw_info["단가"] = 10000
            if raw_info.get("리드타임") == 0:
                raw_info["리드타임"] = 14
            pricing_info = raw_info
            unit_price = pricing_info.get("단가", 10000)
            total_price = unit_price * quantity
            model_name = pricing_info.get("모델명", "N/A")
            supplier = pricing_info.get("공급업체", "미정")
            lead_time_days = pricing_info.get("리드타임", 14)

            # 리드타임 텍스트 변환
            if lead_time_days <= 7:
                lead_time_text = f"{lead_time_days}일"
            else:
                weeks = lead_time_days // 7
                lead_time_text = f"{weeks}주"

            # 클래스 ID 가져오기
            class_id = detections_list[0].get("class_id", -1)
            
            # 연결된 치수 정보 수집
            linked_dimensions = []
            linked_dimension_ids = []
            
            for d in detections_list:
                sym_id = d.get("id")
                if sym_id in symbol_to_dims:
                    for dim in symbol_to_dims[sym_id]:
                        dim_text = dim.get("value", "")
                        # 치수 텍스트 포맷팅 (예: "Ø50", "120")
                        # 값이 승인/수정된 경우 해당 값 사용
                        if dim.get("modified_value"):
                            dim_text = dim.get("modified_value")
                        
                        if dim_text and dim_text not in linked_dimensions:
                             linked_dimensions.append(dim_text)
                        
                        if dim.get("id") not in linked_dimension_ids:
                            linked_dimension_ids.append(dim.get("id"))

            item = {
                "item_no": item_no,
                "class_id": class_id,
                "class_name": class_name,
                "model_name": model_name,
                "quantity": quantity,
                "unit_price": unit_price,
                "total_price": total_price,
                "avg_confidence": round(avg_confidence, 3),
                "detection_ids": [d["id"] for d in detections_list],
                "lead_time": lead_time_text,
                "supplier": supplier,
                "remarks": pricing_info.get("비고", None),
                "dimensions": linked_dimensions,
                "linked_dimension_ids": linked_dimension_ids,
            }
            items.append(item)
            item_no += 1

        # 요약 계산
        total_items = len(items)
        total_quantity = sum(item["quantity"] for item in items)
        subtotal = sum(item["total_price"] for item in items)
        vat = subtotal * 0.1
        total = subtotal + vat

        summary = {
            "total_items": total_items,
            "total_quantity": total_quantity,
            "subtotal": subtotal,
            "vat": vat,
            "total": total,
        }

        bom_data = {
            "session_id": session_id,
            "created_at": datetime.now().isoformat(),
            "items": items,
            "summary": summary,
            "filename": filename,
            "model_id": model_id,
            "detection_count": len(detections),
            "approved_count": len(approved_detections),
        }

        return bom_data

    def export_excel(
        self,
        bom_data: Dict[str, Any],
        output_path: Optional[Path] = None,
        customer_name: Optional[str] = None,
    ) -> Path:
        """BOM을 Excel 파일로 내보내기"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        if output_path is None:
            output_path = self.output_dir / f"bom_{bom_data['session_id']}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        # 스타일 정의
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 제목
        ws.merge_cells('A1:K1')
        ws['A1'] = "부품 명세서 (BOM)"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        # 고객 정보
        if customer_name:
            ws['A3'] = f"고객: {customer_name}"

        ws['A4'] = f"생성일: {bom_data['created_at'][:10]}"
        ws['A5'] = f"파일명: {bom_data.get('filename', 'N/A')}"

        # 헤더
        headers = ["No.", "부품명", "모델명", "수량", "단가", "합계", "치수", "공급업체", "리드타임", "신뢰도", "비고"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # 데이터
        for row_idx, item in enumerate(bom_data["items"], 8):
            ws.cell(row=row_idx, column=1, value=item["item_no"]).border = border
            ws.cell(row=row_idx, column=2, value=item["class_name"]).border = border
            ws.cell(row=row_idx, column=3, value=item.get("model_name", "-")).border = border
            ws.cell(row=row_idx, column=4, value=item["quantity"]).border = border
            ws.cell(row=row_idx, column=5, value=f"₩{item['unit_price']:,}").border = border
            ws.cell(row=row_idx, column=6, value=f"₩{item['total_price']:,}").border = border
            # 치수 정보 (여러 개면 쉼표로 구분)
            dim_text = ", ".join(item.get("dimensions", [])) if item.get("dimensions") else "-"
            ws.cell(row=row_idx, column=7, value=dim_text).border = border
            ws.cell(row=row_idx, column=8, value=item.get("supplier", "-")).border = border
            ws.cell(row=row_idx, column=9, value=item.get("lead_time", "-")).border = border
            ws.cell(row=row_idx, column=10, value=f"{item['avg_confidence']:.1%}").border = border
            ws.cell(row=row_idx, column=11, value=item.get("remarks", "-") or "-").border = border

        # 요약
        summary_row = 8 + len(bom_data["items"]) + 1
        summary = bom_data["summary"]

        ws.cell(row=summary_row, column=5, value="소계").font = header_font
        ws.cell(row=summary_row, column=6, value=f"₩{summary['subtotal']:,.0f}")

        ws.cell(row=summary_row + 1, column=5, value="부가세 (10%)").font = header_font
        ws.cell(row=summary_row + 1, column=6, value=f"₩{summary['vat']:,.0f}")

        ws.cell(row=summary_row + 2, column=5, value="합계").font = Font(bold=True, size=14)
        ws.cell(row=summary_row + 2, column=6, value=f"₩{summary['total']:,.0f}").font = Font(bold=True, size=14)

        # 열 너비 조정
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 20  # 치수
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 12

        wb.save(output_path)
        return output_path

    def export_csv(
        self,
        bom_data: Dict[str, Any],
        output_path: Optional[Path] = None,
    ) -> Path:
        """BOM을 CSV 파일로 내보내기"""
        import csv

        if output_path is None:
            output_path = self.output_dir / f"bom_{bom_data['session_id']}.csv"

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 헤더
            writer.writerow(["No.", "부품명", "모델명", "수량", "단가", "합계", "치수", "공급업체", "리드타임", "신뢰도", "비고"])

            # 데이터
            for item in bom_data["items"]:
                dim_text = ", ".join(item.get("dimensions", [])) if item.get("dimensions") else ""
                writer.writerow([
                    item["item_no"],
                    item["class_name"],
                    item.get("model_name", ""),
                    item["quantity"],
                    item["unit_price"],
                    item["total_price"],
                    dim_text,
                    item.get("supplier", ""),
                    item.get("lead_time", ""),
                    f"{item['avg_confidence']:.1%}",
                    item.get("remarks", "") or "",
                ])

            # 요약
            writer.writerow([])
            summary = bom_data["summary"]
            writer.writerow(["", "", "", "", "소계", summary["subtotal"]])
            writer.writerow(["", "", "", "", "부가세", summary["vat"]])
            writer.writerow(["", "", "", "", "합계", summary["total"]])

        return output_path

    def export_json(
        self,
        bom_data: Dict[str, Any],
        output_path: Optional[Path] = None,
    ) -> Path:
        """BOM을 JSON 파일로 내보내기"""
        import json

        if output_path is None:
            output_path = self.output_dir / f"bom_{bom_data['session_id']}.json"

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(bom_data, f, ensure_ascii=False, indent=2)

        return output_path

    def export_pdf(
        self,
        bom_data: Dict[str, Any],
        output_path: Optional[Path] = None,
        customer_name: Optional[str] = None,
    ) -> Path:
        """BOM을 PDF 파일로 내보내기"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.units import mm
        except ImportError:
            raise NotImplementedError("PDF 내보내기를 위해 reportlab 패키지를 설치해주세요: pip install reportlab")

        if output_path is None:
            output_path = self.output_dir / f"bom_{bom_data['session_id']}.pdf"

        # 한글 폰트 설정 시도
        try:
            # NanumGothic 폰트 시도
            font_paths = [
                "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
                "/usr/share/fonts/nanum/NanumGothic.ttf",
                "/app/fonts/NanumGothic.ttf",
                "NanumGothic.ttf",
            ]
            font_registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('NanumGothic', font_path))
                    font_registered = True
                    break

            if not font_registered:
                logger.warning("한글 폰트를 찾을 수 없습니다. 기본 폰트를 사용합니다.")
        except Exception as e:
            logger.error(f"폰트 등록 실패: {e}")

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )

        elements = []
        styles = getSampleStyleSheet()

        # 제목 스타일
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1  # Center
        )

        # 제목
        elements.append(Paragraph("부품 명세서 (BOM)", title_style))
        elements.append(Spacer(1, 10))

        # 메타 정보
        meta_data = [
            ["생성일", bom_data['created_at'][:10]],
            ["파일명", bom_data.get('filename', 'N/A')],
            ["검출 수", f"{bom_data.get('detection_count', 0)}개"],
            ["승인 수", f"{bom_data.get('approved_count', 0)}개"],
        ]
        if customer_name:
            meta_data.insert(0, ["고객", customer_name])

        meta_table = Table(meta_data, colWidths=[80, 200])
        meta_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 20))

        # BOM 테이블 데이터
        table_data = [["No.", "부품명", "수량", "단가", "합계", "치수", "신뢰도"]]

        for item in bom_data["items"]:
            dim_text = ", ".join(item.get("dimensions", []))[:20] if item.get("dimensions") else "-"
            table_data.append([
                str(item["item_no"]),
                item["class_name"][:25],  # 이름 길이 제한
                str(item["quantity"]),
                f"₩{item['unit_price']:,}",
                f"₩{item['total_price']:,}",
                dim_text,
                f"{item['avg_confidence']:.1%}",
            ])

        # 테이블 생성
        col_widths = [25, 120, 35, 60, 70, 70, 40]
        bom_table = Table(table_data, colWidths=col_widths)
        bom_table.setStyle(TableStyle([
            # 헤더 스타일
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # 데이터 스타일
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),

            # 정렬
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),

            # 그리드
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),

            # 줄무늬 배경
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ]))

        elements.append(bom_table)
        elements.append(Spacer(1, 20))

        # 요약 테이블
        summary = bom_data["summary"]
        summary_data = [
            ["", "소계", f"₩{summary['subtotal']:,.0f}"],
            ["", "부가세 (10%)", f"₩{summary['vat']:,.0f}"],
            ["", "합계", f"₩{summary['total']:,.0f}"],
        ]

        summary_table = Table(summary_data, colWidths=[300, 80, 100])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (1, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (1, -1), (-1, -1), colors.HexColor('#E6F3FF')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(summary_table)

        # PDF 빌드
        doc.build(elements)

        return output_path

    def export(
        self,
        bom_data: Dict[str, Any],
        format: ExportFormat,
        output_path: Optional[Path] = None,
        **kwargs
    ) -> Path:
        """지정된 형식으로 BOM 내보내기"""
        if format == ExportFormat.EXCEL:
            return self.export_excel(bom_data, output_path, **kwargs)
        elif format == ExportFormat.CSV:
            return self.export_csv(bom_data, output_path)
        elif format == ExportFormat.JSON:
            return self.export_json(bom_data, output_path)
        elif format == ExportFormat.PDF:
            return self.export_pdf(bom_data, output_path, **kwargs)
        else:
            raise ValueError(f"지원하지 않는 형식: {format}")

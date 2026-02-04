"""Quotation Service - 견적 집계 서비스

Phase 3: 프로젝트 내 모든 세션 BOM → 재질별 그룹 집계 → PDF/Excel 견적서 내보내기
"""

import json
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List
from collections import defaultdict

from schemas.quotation import (
    SessionQuotationItem,
    MaterialGroup,
    QuotationSummary,
    ProjectQuotationResponse,
    QuotationExportFormat,
    QuotationExportResponse,
)

logger = logging.getLogger(__name__)


class QuotationService:
    """견적 집계 서비스"""

    def __init__(self, data_dir: Path, output_dir: Path):
        self.data_dir = data_dir
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def aggregate_quotation(
        self,
        project_id: str,
        project_service,
        session_service,
    ) -> ProjectQuotationResponse:
        """프로젝트 견적 집계

        1. 프로젝트 정보 조회
        2. 세션 목록 조회
        3. 각 세션에서 견적 항목 구성
        4. 재질별 그룹핑
        5. 요약 계산
        6. quotation.json 저장
        """
        # 1. 프로젝트 정보
        project = project_service.get_project(project_id)
        if not project:
            raise ValueError(f"프로젝트를 찾을 수 없습니다: {project_id}")

        # 2. 세션 목록
        sessions = session_service.list_sessions_by_project(project_id)

        # 3. 각 세션에서 견적 항목 구성
        items: List[SessionQuotationItem] = []
        for session in sessions:
            item = self._build_session_item(session)
            items.append(item)

        # 4. 재질별 그룹핑
        material_groups = self._group_by_material(items)

        # 5. 요약 계산
        summary = self._calculate_summary(items, len(sessions))

        # 6. 응답 구성
        response = ProjectQuotationResponse(
            project_id=project_id,
            project_name=project.get("name", ""),
            customer=project.get("customer", ""),
            created_at=project.get("created_at", datetime.now().isoformat()),
            summary=summary,
            items=items,
            material_groups=material_groups,
        )

        # 7. quotation.json 저장 + 프로젝트 통계 갱신
        self._save_quotation(project_id, response, project_service)

        return response

    def _build_session_item(self, session: Dict[str, Any]) -> SessionQuotationItem:
        """세션 → SessionQuotationItem 변환"""
        metadata = session.get("metadata") or {}
        bom_data = session.get("bom_data") or {}
        bom_summary = bom_data.get("summary") or {}

        return SessionQuotationItem(
            session_id=session.get("session_id", ""),
            drawing_number=metadata.get("drawing_number", ""),
            bom_item_no=metadata.get("bom_item_no", ""),
            description=metadata.get("bom_description", ""),
            material=metadata.get("material", ""),
            bom_quantity=metadata.get("bom_quantity", 1),
            quote_status=metadata.get("quote_status", "pending"),
            bom_item_count=len(bom_data.get("items", [])),
            subtotal=bom_summary.get("subtotal", 0.0),
            vat=bom_summary.get("vat", 0.0),
            total=bom_summary.get("total", 0.0),
            session_status=session.get("status", "uploaded"),
            bom_generated=session.get("bom_generated", False),
        )

    def _group_by_material(
        self, items: List[SessionQuotationItem]
    ) -> List[MaterialGroup]:
        """재질별 defaultdict 그룹핑"""
        groups: Dict[str, List[SessionQuotationItem]] = defaultdict(list)

        for item in items:
            material = item.material or "미지정"
            groups[material].append(item)

        result = []
        for material, group_items in sorted(groups.items()):
            result.append(MaterialGroup(
                material=material,
                item_count=len(group_items),
                total_quantity=sum(i.bom_quantity for i in group_items),
                subtotal=sum(i.subtotal for i in group_items),
                items=group_items,
            ))

        return result

    def _calculate_summary(
        self, items: List[SessionQuotationItem], total_sessions: int
    ) -> QuotationSummary:
        """합계/진행률 계산"""
        completed = sum(1 for i in items if i.bom_generated)
        pending = total_sessions - completed
        quoted = sum(1 for i in items if i.quote_status == "quoted")
        subtotal = sum(i.subtotal for i in items)
        vat = subtotal * 0.1
        total = subtotal + vat
        progress = (completed / total_sessions * 100) if total_sessions > 0 else 0.0

        return QuotationSummary(
            total_sessions=total_sessions,
            completed_sessions=completed,
            pending_sessions=pending,
            quoted_sessions=quoted,
            total_items=len(items),
            subtotal=subtotal,
            vat=vat,
            total=total,
            progress_percent=round(progress, 1),
        )

    def _save_quotation(
        self,
        project_id: str,
        data: ProjectQuotationResponse,
        project_service,
    ):
        """quotation.json 저장, project 통계 갱신"""
        project_dir = project_service.projects_dir / project_id
        project_dir.mkdir(parents=True, exist_ok=True)

        quotation_file = project_dir / "quotation.json"
        with open(quotation_file, "w", encoding="utf-8") as f:
            json.dump(data.model_dump(), f, ensure_ascii=False, indent=2, default=str)

        # 프로젝트 통계 업데이트
        project = project_service.get_project(project_id)
        if project:
            project["quoted_count"] = data.summary.quoted_sessions
            project["total_quotation"] = data.summary.total
            project["updated_at"] = datetime.now().isoformat()
            project_service._save_project(project_id)

        logger.info(f"견적 집계 저장: {project_id} → {quotation_file}")

    def _load_quotation(
        self, project_id: str, project_service
    ) -> Optional[ProjectQuotationResponse]:
        """캐시된 quotation.json 로드"""
        project_dir = project_service.projects_dir / project_id
        quotation_file = project_dir / "quotation.json"

        if not quotation_file.exists():
            return None

        try:
            with open(quotation_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            return ProjectQuotationResponse(**data)
        except Exception as e:
            logger.error(f"견적 데이터 로드 실패: {e}")
            return None

    def export_pdf(
        self,
        quotation_data: ProjectQuotationResponse,
        customer_name: Optional[str] = None,
        include_material_breakdown: bool = True,
        notes: Optional[str] = None,
    ) -> Path:
        """reportlab 기반 견적서 PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.units import mm
        except ImportError:
            raise NotImplementedError(
                "PDF 내보내기를 위해 reportlab 패키지를 설치해주세요: pip install reportlab"
            )

        output_path = self.output_dir / f"quotation_{quotation_data.project_id}.pdf"

        # 한글 폰트 설정
        font_paths = [
            "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
            "/usr/share/fonts/nanum/NanumGothic.ttf",
            "/app/fonts/NanumGothic.ttf",
            "NanumGothic.ttf",
        ]
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('NanumGothic', font_path))
                except Exception:
                    pass
                break

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,
        )

        # 제목
        elements.append(Paragraph("견적서 (Quotation)", title_style))
        elements.append(Spacer(1, 10))

        # 메타 정보
        effective_customer = customer_name or quotation_data.customer
        meta_data = [
            ["프로젝트", quotation_data.project_name],
            ["고객", effective_customer],
            ["생성일", quotation_data.created_at[:10]],
            ["세션 수", f"{quotation_data.summary.total_sessions}개"],
            ["완료", f"{quotation_data.summary.completed_sessions}개"],
        ]
        meta_table = Table(meta_data, colWidths=[80, 200])
        meta_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 20))

        # 세션별 견적 테이블
        table_data = [["No.", "도면번호", "재질", "수량", "소계", "상태"]]
        for idx, item in enumerate(quotation_data.items, 1):
            status = "완료" if item.bom_generated else "대기"
            table_data.append([
                str(idx),
                item.drawing_number[:25] or "-",
                item.material or "-",
                str(item.bom_quantity),
                f"₩{item.subtotal:,.0f}" if item.subtotal > 0 else "-",
                status,
            ])

        col_widths = [30, 120, 60, 40, 80, 40]
        main_table = Table(table_data, colWidths=col_widths)
        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#F0F0F0')]),
        ]))
        elements.append(main_table)
        elements.append(Spacer(1, 15))

        # 재질별 분류 (선택)
        if include_material_breakdown and quotation_data.material_groups:
            elements.append(Spacer(1, 10))
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=10,
            )
            elements.append(Paragraph("재질별 분류", subtitle_style))

            mat_data = [["재질", "품목수", "총 수량", "소계"]]
            for mg in quotation_data.material_groups:
                mat_data.append([
                    mg.material,
                    str(mg.item_count),
                    str(mg.total_quantity),
                    f"₩{mg.subtotal:,.0f}" if mg.subtotal > 0 else "-",
                ])

            mat_table = Table(mat_data, colWidths=[100, 60, 60, 100])
            mat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B9BD5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(mat_table)
            elements.append(Spacer(1, 15))

        # 요약 테이블
        summary = quotation_data.summary
        summary_data = [
            ["", "소계", f"₩{summary.subtotal:,.0f}"],
            ["", "부가세 (10%)", f"₩{summary.vat:,.0f}"],
            ["", "합계", f"₩{summary.total:,.0f}"],
        ]
        summary_table = Table(summary_data, colWidths=[200, 80, 100])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (1, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (1, -1), (-1, -1), colors.HexColor('#E6F3FF')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)

        # 비고
        if notes:
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"비고: {notes}", styles['Normal']))

        doc.build(elements)
        return output_path

    def export_excel(
        self,
        quotation_data: ProjectQuotationResponse,
        customer_name: Optional[str] = None,
        include_material_breakdown: bool = True,
        notes: Optional[str] = None,
    ) -> Path:
        """openpyxl 기반 견적서 Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        output_path = self.output_dir / f"quotation_{quotation_data.project_id}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "견적서"

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # 제목
        ws.merge_cells('A1:F1')
        ws['A1'] = "견적서 (Quotation)"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        # 메타 정보
        effective_customer = customer_name or quotation_data.customer
        ws['A3'] = f"프로젝트: {quotation_data.project_name}"
        ws['A4'] = f"고객: {effective_customer}"
        ws['A5'] = f"생성일: {quotation_data.created_at[:10]}"

        # 헤더
        headers = ["No.", "도면번호", "설명", "재질", "수량", "소계"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # 데이터
        for row_idx, item in enumerate(quotation_data.items, 8):
            ws.cell(row=row_idx, column=1, value=row_idx - 7).border = border
            ws.cell(row=row_idx, column=2,
                    value=item.drawing_number or "-").border = border
            ws.cell(row=row_idx, column=3,
                    value=item.description or "-").border = border
            ws.cell(row=row_idx, column=4,
                    value=item.material or "-").border = border
            ws.cell(row=row_idx, column=5,
                    value=item.bom_quantity).border = border
            ws.cell(row=row_idx, column=6,
                    value=f"₩{item.subtotal:,.0f}" if item.subtotal else "-"
                    ).border = border

        # 요약
        summary_row = 8 + len(quotation_data.items) + 1
        summary = quotation_data.summary
        bold_font = Font(bold=True, size=12)

        ws.cell(row=summary_row, column=5, value="소계").font = bold_font
        ws.cell(row=summary_row, column=6, value=f"₩{summary.subtotal:,.0f}")
        ws.cell(row=summary_row + 1, column=5, value="부가세 (10%)").font = bold_font
        ws.cell(row=summary_row + 1, column=6, value=f"₩{summary.vat:,.0f}")
        ws.cell(row=summary_row + 2, column=5, value="합계").font = Font(bold=True, size=14)
        ws.cell(row=summary_row + 2, column=6, value=f"₩{summary.total:,.0f}"
                ).font = Font(bold=True, size=14)

        # 재질별 분류 시트
        if include_material_breakdown and quotation_data.material_groups:
            ws2 = wb.create_sheet(title="재질별 분류")
            mat_headers = ["재질", "품목수", "총 수량", "소계"]
            for col, header in enumerate(mat_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            for row_idx, mg in enumerate(quotation_data.material_groups, 2):
                ws2.cell(row=row_idx, column=1, value=mg.material).border = border
                ws2.cell(row=row_idx, column=2, value=mg.item_count).border = border
                ws2.cell(row=row_idx, column=3, value=mg.total_quantity).border = border
                ws2.cell(row=row_idx, column=4,
                         value=f"₩{mg.subtotal:,.0f}" if mg.subtotal else "-"
                         ).border = border

        # 열 너비
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14

        wb.save(output_path)
        return output_path

    def export(
        self,
        quotation_data: ProjectQuotationResponse,
        format: QuotationExportFormat,
        customer_name: Optional[str] = None,
        include_material_breakdown: bool = True,
        notes: Optional[str] = None,
    ) -> QuotationExportResponse:
        """format 디스패치"""
        if format == QuotationExportFormat.PDF:
            path = self.export_pdf(
                quotation_data, customer_name, include_material_breakdown, notes
            )
        elif format == QuotationExportFormat.EXCEL:
            path = self.export_excel(
                quotation_data, customer_name, include_material_breakdown, notes
            )
        else:
            raise ValueError(f"지원하지 않는 형식: {format}")

        file_size = path.stat().st_size

        return QuotationExportResponse(
            project_id=quotation_data.project_id,
            format=format,
            filename=path.name,
            file_path=str(path),
            file_size=file_size,
            created_at=datetime.now().isoformat(),
        )


# 싱글톤 인스턴스
_quotation_service: Optional[QuotationService] = None


def get_quotation_service(
    data_dir: Optional[Path] = None,
    output_dir: Optional[Path] = None,
) -> QuotationService:
    """QuotationService 싱글톤 인스턴스 반환"""
    global _quotation_service

    if _quotation_service is None:
        if data_dir is None:
            data_dir = Path("/app/data")
        if output_dir is None:
            output_dir = Path("/app/data/exports")
        _quotation_service = QuotationService(data_dir, output_dir)

    return _quotation_service

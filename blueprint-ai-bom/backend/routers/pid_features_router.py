"""P&ID 분석 기능 라우터

P&ID 도면 분석을 위한 Human-in-the-Loop 워크플로우 API

기능:
- Valve Detection: P&ID 밸브 태그 검출
- Equipment Detection: 장비 태그 검출 및 목록화
- Design Checklist: 설계 규칙 검증
- Deviation Analysis: 기준 대비 편차 분석 (향후)

워크플로우:
1. 세션 생성 (features: pid_valve_detection, pid_equipment_detection, pid_design_checklist)
2. 각 기능별 검출 실행
3. 검증 큐에서 검토/승인/수정
4. Excel 내보내기
"""

import os
import uuid
import time
import logging
from io import BytesIO
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
import httpx

from schemas.pid_features import (
    PIDFeature,
    VerificationStatus,
    ValveCategory,
    ValveItem,
    ValveDetectionResponse,
    EquipmentType,
    EquipmentItem,
    EquipmentDetectionResponse,
    ChecklistStatus,
    ChecklistItem,
    ChecklistResponse,
    PIDVerifyRequest,
    PIDBulkVerifyRequest,
    PIDVerificationQueue,
    PIDExportResponse,
)
from services.active_learning_service import active_learning_service

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/pid-features", tags=["P&ID Features"])

# =====================
# Configuration
# =====================

PID_ANALYZER_URL = os.getenv("PID_ANALYZER_URL", "http://pid-analyzer-api:5018")
DESIGN_CHECKER_URL = os.getenv("DESIGN_CHECKER_URL", "http://design-checker-api:5019")
PADDLEOCR_URL = os.getenv("PADDLEOCR_URL", "http://paddleocr-api:5006")
LINE_DETECTOR_URL = os.getenv("LINE_DETECTOR_URL", "http://line-detector-api:5016")

# 서비스 의존성
_session_service = None


def set_pid_features_service(session_service):
    """서비스 주입"""
    global _session_service
    _session_service = session_service


def get_session_service():
    if _session_service is None:
        raise HTTPException(status_code=500, detail="Session service not initialized")
    return _session_service


# =====================
# Valve Detection
# =====================

@router.post("/{session_id}/valve/detect", response_model=ValveDetectionResponse)
async def detect_valves(
    session_id: str,
    rule_id: str = Query(default="valve_detection_default", description="추출 규칙 ID"),
    profile: str = Query(default="default", description="검출 프로필 (default, bwms, chemical 등)"),
    language: str = Query(default="en", description="OCR 언어")
):
    """
    P&ID 밸브 검출

    P&ID 이미지에서 밸브 태그를 추출합니다.
    프로필에 따라 다른 검출 규칙이 적용됩니다.
    """
    start_time = time.time()
    session_service = get_session_service()

    session = session_service.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    file_path = session.get("file_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="이미지 파일을 찾을 수 없습니다")

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            with open(file_path, "rb") as f:
                files = {"file": (os.path.basename(file_path), f, "image/png")}
                data = {"rule_id": rule_id, "language": language}

                response = await client.post(
                    f"{PID_ANALYZER_URL}/api/v1/valve-signal/extract",
                    files=files,
                    data=data
                )

        if response.status_code != 200:
            logger.error(f"PID Analyzer error: {response.status_code} - {response.text}")
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Valve detection failed: {response.text}"
            )

        result = response.json()
        extracted_items = result.get("data", {}).get("all_extracted_items", [])

        valves = []
        category_counts = {cat.value: 0 for cat in ValveCategory}

        for idx, item in enumerate(extracted_items):
            valve_id = item.get("id") or item.get("matched_text") or item.get("text", "")
            valve_type = valve_id.split("-")[0] if "-" in valve_id else "Unknown"
            category = _classify_valve_category(valve_id, valve_type)
            category_counts[category.value] += 1

            raw_bbox = item.get("bbox")
            converted_bbox = None
            if raw_bbox:
                try:
                    if isinstance(raw_bbox[0], (list, tuple)):
                        xs = [p[0] for p in raw_bbox]
                        ys = [p[1] for p in raw_bbox]
                        converted_bbox = [float(min(xs)), float(min(ys)), float(max(xs)), float(max(ys))]
                    else:
                        converted_bbox = [float(v) for v in raw_bbox[:4]]
                except Exception:
                    converted_bbox = None

            valve = ValveItem(
                id=f"valve_{session_id[:8]}_{idx}",
                valve_id=valve_id,
                valve_type=valve_type,
                category=category,
                region_name=item.get("region_name", ""),
                confidence=item.get("confidence", 0.0),
                bbox=converted_bbox,
                verification_status=VerificationStatus.PENDING
            )
            valves.append(valve)

        session_service.update_session(session_id, {
            "pid_valves": [v.model_dump() for v in valves],
            "pid_valve_count": len(valves)
        })

        processing_time = time.time() - start_time

        return ValveDetectionResponse(
            session_id=session_id,
            valves=valves,
            total_count=len(valves),
            categories=category_counts,
            processing_time=round(processing_time, 3),
            regions_detected=result.get("data", {}).get("line_detector", {}).get("total_regions", 0)
        )

    except httpx.RequestError as e:
        logger.error(f"PID Analyzer connection error: {e}")
        raise HTTPException(status_code=503, detail=f"PID Analyzer 서비스 연결 실패: {e}")


def _classify_valve_category(valve_id: str, valve_type: str) -> ValveCategory:
    """밸브 카테고리 분류"""
    valve_id_upper = valve_id.upper()
    valve_type_upper = valve_type.upper()

    # 제어 밸브
    if valve_type_upper in ["CV", "FCV", "PCV", "TCV", "LCV"]:
        return ValveCategory.CONTROL

    # 차단 밸브
    if valve_type_upper in ["XV", "BV", "GV", "IV"]:
        return ValveCategory.ISOLATION

    # 안전 밸브
    if valve_type_upper in ["SV", "PSV", "TSV"]:
        return ValveCategory.SAFETY

    # 체크 밸브
    if valve_type_upper in ["CHK", "CK", "NRV"]:
        return ValveCategory.CHECK

    # 릴리프 밸브
    if valve_type_upper in ["RV", "PRV"]:
        return ValveCategory.RELIEF

    return ValveCategory.UNKNOWN


# =====================
# Equipment Detection
# =====================

@router.post("/{session_id}/equipment/detect", response_model=EquipmentDetectionResponse)
async def detect_equipment(
    session_id: str,
    profile: str = Query(default="default", description="장비 프로필"),
    language: str = Query(default="en", description="OCR 언어")
):
    """
    P&ID 장비 검출

    P&ID 이미지에서 장비 태그를 검출합니다.
    프로필에 따라 특화된 장비 타입을 인식합니다.
    """
    start_time = time.time()
    session_service = get_session_service()

    session = session_service.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    file_path = session.get("file_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="이미지 파일을 찾을 수 없습니다")

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            with open(file_path, "rb") as f:
                files = {"file": (os.path.basename(file_path), f, "image/png")}
                data = {"profile_id": profile, "language": language}

                response = await client.post(
                    f"{PID_ANALYZER_URL}/api/v1/equipment/detect",
                    files=files,
                    data=data
                )

        if response.status_code != 200:
            logger.error(f"Equipment detection error: {response.status_code} - {response.text}")
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Equipment detection failed: {response.text}"
            )

        result = response.json()
        detected_equipment = result.get("data", {}).get("equipment", [])

        equipment_items = []
        type_counts = {et.value: 0 for et in EquipmentType}
        vendor_supply_count = 0

        for idx, eq in enumerate(detected_equipment):
            tag = eq.get("tag", "")
            eq_type = _classify_equipment_type(tag)
            type_counts[eq_type.value] += 1

            is_vendor_supply = eq.get("vendor_supply", False) or "*" in tag
            if is_vendor_supply:
                vendor_supply_count += 1

            equipment_item = EquipmentItem(
                id=f"eq_{session_id[:8]}_{idx}",
                tag=tag,
                equipment_type=eq_type,
                description=eq.get("description", ""),
                vendor_supply=is_vendor_supply,
                confidence=eq.get("confidence", 0.0),
                bbox=eq.get("bbox"),
                verification_status=VerificationStatus.PENDING
            )
            equipment_items.append(equipment_item)

        session_service.update_session(session_id, {
            "pid_equipment": [eq.model_dump() for eq in equipment_items],
            "pid_equipment_count": len(equipment_items)
        })

        processing_time = time.time() - start_time

        return EquipmentDetectionResponse(
            session_id=session_id,
            equipment=equipment_items,
            total_count=len(equipment_items),
            by_type=type_counts,
            vendor_supply_count=vendor_supply_count,
            processing_time=round(processing_time, 3)
        )

    except httpx.RequestError as e:
        logger.error(f"PID Analyzer connection error: {e}")
        raise HTTPException(status_code=503, detail=f"PID Analyzer 서비스 연결 실패: {e}")


def _classify_equipment_type(tag: str) -> EquipmentType:
    """장비 타입 분류"""
    tag_upper = tag.upper()

    if any(p in tag_upper for p in ["PUMP", "P-", "PP-"]):
        return EquipmentType.PUMP
    if any(p in tag_upper for p in ["VALVE", "V-", "XV", "SV", "CV"]):
        return EquipmentType.VALVE
    if any(p in tag_upper for p in ["TANK", "T-", "TK"]):
        return EquipmentType.TANK
    if any(p in tag_upper for p in ["HX", "E-", "EXCHANGER"]):
        return EquipmentType.HEAT_EXCHANGER
    if any(p in tag_upper for p in ["COMP", "C-", "K-"]):
        return EquipmentType.COMPRESSOR
    if any(p in tag_upper for p in ["FILTER", "F-", "FLT"]):
        return EquipmentType.FILTER
    if any(p in tag_upper for p in ["CTRL", "PLC", "DCS"]):
        return EquipmentType.CONTROLLER
    if any(p in tag_upper for p in ["PT", "TT", "FT", "LT", "AT"]):
        return EquipmentType.SENSOR

    return EquipmentType.OTHER


# =====================
# Design Checklist
# =====================

@router.post("/{session_id}/checklist/check", response_model=ChecklistResponse)
async def check_design_checklist(
    session_id: str,
    rule_profile: str = Query(default="default", description="규칙 프로필 (default, bwms, chemical 등)"),
    enabled_rules: Optional[str] = Query(default=None, description="활성화할 규칙 (쉼표 구분)")
):
    """
    설계 체크리스트 검사

    P&ID 도면에 대해 설계 규칙을 자동 검증합니다.
    프로필에 따라 다른 규칙 세트가 적용됩니다.
    """
    start_time = time.time()
    session_service = get_session_service()

    session = session_service.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    symbols = session.get("detections", [])
    connections = session.get("connections", [])
    texts = session.get("ocr_texts", [])

    if not symbols:
        logger.warning(f"Session {session_id}: No symbols detected, checklist may be incomplete")

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            import json
            data = {
                "symbols": json.dumps(symbols),
                "connections": json.dumps(connections),
                "texts": json.dumps(texts),
                "enabled_rules": enabled_rules or "",
                "rule_profile": rule_profile
            }

            response = await client.post(
                f"{DESIGN_CHECKER_URL}/api/v1/check/bwms",
                data=data
            )

        if response.status_code != 200:
            logger.error(f"Design Checker error: {response.status_code} - {response.text}")
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Checklist check failed: {response.text}"
            )

        result = response.json()
        violations = result.get("data", {}).get("violations", [])

        checklist_items = []
        summary = {
            "Pass": 0,
            "Fail": 0,
            "N/A": 0,
            "Pending": 0,
            "Manual Required": 0
        }

        for idx, violation in enumerate(violations):
            status = ChecklistStatus.FAIL if violation.get("severity") == "error" else ChecklistStatus.PENDING
            summary[status.value] += 1

            checklist_item = ChecklistItem(
                id=f"check_{session_id[:8]}_{idx}",
                item_no=idx + 1,
                category=violation.get("category", "design"),
                description=violation.get("description", ""),
                description_ko=violation.get("rule_name", ""),
                auto_status=status,
                final_status=status,
                evidence=str(violation.get("affected_elements", [])),
                confidence=0.8 if status == ChecklistStatus.FAIL else 0.5,
                verification_status=VerificationStatus.PENDING
            )
            checklist_items.append(checklist_item)

        session_service.update_session(session_id, {
            "pid_checklist_items": [ci.model_dump() for ci in checklist_items],
            "pid_checklist_count": len(checklist_items)
        })

        processing_time = time.time() - start_time

        total = len(checklist_items)
        compliance_rate = (summary["Pass"] / total * 100) if total > 0 else 0.0

        return ChecklistResponse(
            session_id=session_id,
            items=checklist_items,
            total_count=len(checklist_items),
            summary=summary,
            compliance_rate=round(compliance_rate, 2),
            processing_time=round(processing_time, 3)
        )

    except httpx.RequestError as e:
        logger.error(f"Design Checker connection error: {e}")
        raise HTTPException(status_code=503, detail=f"Design Checker 서비스 연결 실패: {e}")


# =====================
# Deviation Analysis
# =====================

# Import deviation schemas at the top of the file are handled by the existing imports

@router.post("/{session_id}/deviation/analyze")
async def analyze_deviations(
    session_id: str,
    request: Optional[Dict[str, Any]] = None
):
    """
    편차 분석 실행

    확장 가능한 분석 패턴:
    - revision_compare: 리비전 비교 (baseline_session_id 필요)
    - standard_check: 표준 검사 (ISO 10628, ISA 5.1 등)
    - design_spec_check: 설계 기준서 대비 검사
    - vlm_analysis: VLM 기반 분석

    향후 요구사항에 따라 각 분석 타입의 상세 로직 구현 가능
    """
    from schemas.pid_features import (
        DeviationCategory,
        DeviationSeverity,
        DeviationSource,
        DeviationItem,
    )

    start_time = time.time()
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    # 요청 파라미터 파싱
    request = request or {}
    analysis_types = request.get("analysis_types", ["standard_check"])
    baseline_session_id = request.get("baseline_session_id")
    standards = request.get("standards", ["ISO_10628"])
    severity_threshold = request.get("severity_threshold", "low")
    include_info = request.get("include_info", True)
    profile = request.get("profile", "default")

    deviations = []
    standards_applied = []

    # 분석 타입별 처리 (확장 포인트)
    for analysis_type in analysis_types:

        if analysis_type == "revision_compare" and baseline_session_id:
            # 리비전 비교 분석 (기존 longterm_router의 revision comparison 활용 가능)
            logger.info(f"[Deviation] Revision compare: {session_id} vs {baseline_session_id}")

            # TODO: 실제 구현 시 longterm_router.compare_revisions 호출
            # 현재는 플레이스홀더 - 리비전 비교 결과를 DeviationItem으로 변환
            deviations.append(DeviationItem(
                id=f"dev_rev_{session_id[:8]}_001",
                category=DeviationCategory.REVISION_MODIFIED,
                severity=DeviationSeverity.INFO,
                source=DeviationSource.REVISION_COMPARE,
                title="리비전 비교 대기 중",
                description="리비전 비교 분석을 위해 베이스라인 세션이 설정되었습니다.",
                baseline_info=baseline_session_id,
                confidence=1.0,
                verification_status=VerificationStatus.PENDING
            ))

        elif analysis_type == "standard_check":
            # 표준 검사 (Design Checker API 활용)
            standards_applied.extend(standards)
            logger.info(f"[Deviation] Standard check with standards: {standards}")

            # Design Checker API 호출하여 표준 위반 사항 검출
            try:
                import json as json_module
                async with httpx.AsyncClient(timeout=60.0) as client:
                    # Design Checker는 Form 데이터를 기대함
                    symbols_data = session.get("detections", [])
                    lines_data = session.get("lines", [])
                    texts_data = session.get("ocr_results", [])

                    form_data = {
                        "symbols": json_module.dumps(symbols_data),
                        "connections": json_module.dumps([]),
                        "lines": json_module.dumps(lines_data),
                        "texts": json_module.dumps(texts_data),
                        "enabled_rules": ""  # 전체 규칙
                    }

                    response = await client.post(
                        f"{DESIGN_CHECKER_URL}/api/v1/check/bwms",
                        data=form_data
                    )

                    if response.status_code == 200:
                        result = response.json()
                        violations = result.get("data", {}).get("violations", [])

                        for idx, violation in enumerate(violations):
                            severity_map = {
                                "error": DeviationSeverity.HIGH,
                                "warning": DeviationSeverity.MEDIUM,
                                "info": DeviationSeverity.LOW
                            }
                            dev_severity = severity_map.get(
                                violation.get("severity", "warning"),
                                DeviationSeverity.MEDIUM
                            )

                            deviations.append(DeviationItem(
                                id=f"dev_std_{session_id[:8]}_{idx:03d}",
                                category=DeviationCategory.STANDARD_VIOLATION,
                                severity=dev_severity,
                                source=DeviationSource.STANDARD_CHECK,
                                title=violation.get("rule_name", "표준 위반"),
                                description=violation.get("description", ""),
                                location=str(violation.get("affected_elements", [])),
                                reference_standard=violation.get("rule_id", ""),
                                confidence=0.8,
                                action_required=violation.get("recommendation", "검토 필요"),
                                verification_status=VerificationStatus.PENDING
                            ))

            except httpx.RequestError as e:
                logger.warning(f"Design Checker connection failed: {e}")
                # 연결 실패 시 플레이스홀더 반환
                deviations.append(DeviationItem(
                    id=f"dev_std_{session_id[:8]}_err",
                    category=DeviationCategory.OTHER,
                    severity=DeviationSeverity.INFO,
                    source=DeviationSource.STANDARD_CHECK,
                    title="표준 검사 서비스 연결 실패",
                    description=f"Design Checker API 연결에 실패했습니다: {str(e)}",
                    confidence=1.0,
                    verification_status=VerificationStatus.PENDING
                ))

        elif analysis_type == "design_spec_check":
            # 설계 기준서 대비 검사 (향후 구현)
            logger.info(f"[Deviation] Design spec check (placeholder)")
            deviations.append(DeviationItem(
                id=f"dev_spec_{session_id[:8]}_001",
                category=DeviationCategory.DESIGN_MISSING,
                severity=DeviationSeverity.INFO,
                source=DeviationSource.DESIGN_SPEC_CHECK,
                title="설계 기준서 검사 대기 중",
                description="설계 기준서 ID가 설정되면 상세 검사가 진행됩니다.",
                confidence=1.0,
                verification_status=VerificationStatus.PENDING
            ))

        elif analysis_type == "vlm_analysis":
            # VLM 기반 분석 (향후 구현)
            logger.info(f"[Deviation] VLM analysis (placeholder)")
            deviations.append(DeviationItem(
                id=f"dev_vlm_{session_id[:8]}_001",
                category=DeviationCategory.OTHER,
                severity=DeviationSeverity.INFO,
                source=DeviationSource.VLM_ANALYSIS,
                title="VLM 분석 대기 중",
                description="Vision-Language Model 기반 편차 분석이 예정되어 있습니다.",
                confidence=1.0,
                verification_status=VerificationStatus.PENDING
            ))

    # 심각도 필터링
    severity_order = ["critical", "high", "medium", "low", "info"]
    threshold_idx = severity_order.index(severity_threshold.lower()) if severity_threshold.lower() in severity_order else 4

    if not include_info and threshold_idx == 4:
        threshold_idx = 3  # info 제외

    filtered_deviations = [
        d for d in deviations
        if severity_order.index(d.severity.value) <= threshold_idx
    ]

    # 세션에 저장
    session_service.update_session(session_id, {
        "pid_deviations": [d.model_dump() for d in filtered_deviations],
        "pid_deviation_count": len(filtered_deviations)
    })

    processing_time = time.time() - start_time

    # 통계 계산
    by_category = {}
    by_severity = {}
    by_source = {}
    for d in filtered_deviations:
        by_category[d.category.value] = by_category.get(d.category.value, 0) + 1
        by_severity[d.severity.value] = by_severity.get(d.severity.value, 0) + 1
        by_source[d.source.value] = by_source.get(d.source.value, 0) + 1

    return {
        "session_id": session_id,
        "deviations": [d.model_dump() for d in filtered_deviations],
        "total_count": len(filtered_deviations),
        "by_category": by_category,
        "by_severity": by_severity,
        "by_source": by_source,
        "analysis_types_used": analysis_types,
        "standards_applied": standards_applied,
        "baseline_session_id": baseline_session_id,
        "processing_time": round(processing_time, 3)
    }


@router.get("/{session_id}/deviation/list")
async def get_deviations(
    session_id: str,
    category: Optional[str] = Query(None, description="편차 카테고리 필터"),
    severity: Optional[str] = Query(None, description="심각도 필터"),
    source: Optional[str] = Query(None, description="소스 필터")
):
    """
    편차 목록 조회 (필터링 지원)
    """
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    deviations = session.get("pid_deviations", [])

    # 필터링
    if category:
        deviations = [d for d in deviations if d.get("category") == category]
    if severity:
        deviations = [d for d in deviations if d.get("severity") == severity]
    if source:
        deviations = [d for d in deviations if d.get("source") == source]

    return {
        "session_id": session_id,
        "deviations": deviations,
        "total_count": len(deviations),
        "filters_applied": {
            "category": category,
            "severity": severity,
            "source": source
        }
    }


# =====================
# Verification (공통)
# =====================

@router.get("/{session_id}/verify/queue", response_model=PIDVerificationQueue)
async def get_verification_queue(
    session_id: str,
    item_type: str = Query(..., description="항목 타입 (valve, equipment, checklist_item)")
):
    """
    검증 큐 조회

    우선순위 순으로 정렬된 검증 항목 반환
    """
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    field_map = {
        "valve": "pid_valves",
        "equipment": "pid_equipment",
        "checklist_item": "pid_checklist_items",
        "deviation": "pid_deviations"
    }

    field_name = field_map.get(item_type)
    if not field_name:
        raise HTTPException(status_code=400, detail=f"Invalid item_type: {item_type}")

    items = session.get(field_name, [])
    queue = active_learning_service.get_verification_queue(items, item_type)
    stats = active_learning_service.get_verification_stats(items, item_type)

    return PIDVerificationQueue(
        session_id=session_id,
        item_type=item_type,
        queue=[item.to_dict() for item in queue],
        stats=stats,
        thresholds=active_learning_service.thresholds
    )


@router.post("/{session_id}/verify")
async def verify_item(
    session_id: str,
    request: PIDVerifyRequest
):
    """
    개별 항목 검증
    """
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    field_map = {
        "valve": "pid_valves",
        "equipment": "pid_equipment",
        "checklist_item": "pid_checklist_items",
        "deviation": "pid_deviations"
    }

    field_name = field_map.get(request.item_type)
    if not field_name:
        raise HTTPException(status_code=400, detail=f"Invalid item_type: {request.item_type}")

    items = session.get(field_name, [])
    item_found = False

    for item in items:
        if item.get("id") == request.item_id:
            item_found = True

            if request.action == "approve":
                item["verification_status"] = VerificationStatus.APPROVED.value
            elif request.action == "reject":
                item["verification_status"] = VerificationStatus.REJECTED.value
            elif request.action == "modify":
                item["verification_status"] = VerificationStatus.MODIFIED.value
                if request.modified_data:
                    item.update(request.modified_data)

            item["modified_at"] = datetime.now().isoformat()
            item["notes"] = request.notes
            break

    if not item_found:
        raise HTTPException(status_code=404, detail=f"Item not found: {request.item_id}")

    session_service.update_session(session_id, {field_name: items})

    active_learning_service.log_verification(
        item_id=request.item_id,
        item_type=request.item_type,
        original_data={},
        user_action=request.action,
        modified_data=request.modified_data,
        session_id=session_id
    )

    return {
        "status": "success",
        "session_id": session_id,
        "item_id": request.item_id,
        "action": request.action
    }


@router.post("/{session_id}/verify/bulk")
async def bulk_verify_items(
    session_id: str,
    request: PIDBulkVerifyRequest
):
    """
    일괄 검증
    """
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    field_map = {
        "valve": "pid_valves",
        "equipment": "pid_equipment",
        "checklist_item": "pid_checklist_items",
        "deviation": "pid_deviations"
    }

    field_name = field_map.get(request.item_type)
    if not field_name:
        raise HTTPException(status_code=400, detail=f"Invalid item_type: {request.item_type}")

    items = session.get(field_name, [])
    updated_count = 0

    for item in items:
        if item.get("id") in request.item_ids:
            if request.action == "approve":
                item["verification_status"] = VerificationStatus.APPROVED.value
            elif request.action == "reject":
                item["verification_status"] = VerificationStatus.REJECTED.value

            item["modified_at"] = datetime.now().isoformat()
            updated_count += 1

    session_service.update_session(session_id, {field_name: items})

    return {
        "status": "success",
        "session_id": session_id,
        "updated_count": updated_count,
        "action": request.action
    }


# =====================
# Export
# =====================

@router.post("/{session_id}/export")
async def export_to_excel(
    session_id: str,
    export_type: str = Query(..., description="내보내기 타입 (valve, equipment, checklist, all)"),
    project_name: str = Query(default="Unknown Project", description="프로젝트명"),
    drawing_no: str = Query(default="N/A", description="도면 번호"),
    include_rejected: bool = Query(default=False, description="거부된 항목 포함")
):
    """
    Excel 내보내기
    """
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def apply_header_style(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    sheets_created = 0

    # Valve 시트
    if export_type in ["valve", "all"]:
        valves = session.get("pid_valves", [])
        if not include_rejected:
            valves = [v for v in valves if v.get("verification_status") != "rejected"]

        if valves:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Valve List"

            ws.append(["No", "Valve ID", "Type", "Category", "Region", "Confidence", "Status", "Notes"])
            apply_header_style(ws)

            for idx, v in enumerate(valves, 1):
                ws.append([
                    idx,
                    v.get("valve_id", ""),
                    v.get("valve_type", ""),
                    v.get("category", ""),
                    v.get("region_name", ""),
                    round(v.get("confidence", 0), 3),
                    v.get("verification_status", ""),
                    v.get("notes", "")
                ])

            sheets_created += 1

    # Equipment 시트
    if export_type in ["equipment", "all"]:
        equipment = session.get("pid_equipment", [])
        if not include_rejected:
            equipment = [e for e in equipment if e.get("verification_status") != "rejected"]

        if equipment:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Equipment List"

            ws.append(["No", "Tag", "Type", "Description", "Vendor Supply", "Confidence", "Status", "Notes"])
            apply_header_style(ws)

            for idx, e in enumerate(equipment, 1):
                ws.append([
                    idx,
                    e.get("tag", ""),
                    e.get("equipment_type", ""),
                    e.get("description", ""),
                    "Yes" if e.get("vendor_supply") else "No",
                    round(e.get("confidence", 0), 3),
                    e.get("verification_status", ""),
                    e.get("notes", "")
                ])

            sheets_created += 1

    # Checklist 시트
    if export_type in ["checklist", "all"]:
        checklist_items = session.get("pid_checklist_items", [])
        if not include_rejected:
            checklist_items = [c for c in checklist_items if c.get("verification_status") != "rejected"]

        if checklist_items:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Design Checklist"

            ws.append(["No", "Category", "Description", "Auto Status", "Final Status", "Evidence", "Reviewer Notes"])
            apply_header_style(ws)

            for c in checklist_items:
                ws.append([
                    c.get("item_no", ""),
                    c.get("category", ""),
                    c.get("description", ""),
                    c.get("auto_status", ""),
                    c.get("final_status", ""),
                    c.get("evidence", ""),
                    c.get("reviewer_notes", "")
                ])

            sheets_created += 1

    # Deviation 시트
    if export_type in ["deviation", "all"]:
        deviations = session.get("pid_deviations", [])
        if not include_rejected:
            deviations = [d for d in deviations if d.get("verification_status") != "rejected"]

        if deviations:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Deviation List"

            ws.append([
                "No", "Category", "Severity", "Source", "Title", "Description",
                "Location", "Reference Standard", "Reference Value", "Actual Value",
                "Action Required", "Action Taken", "Status", "Notes"
            ])
            apply_header_style(ws)

            for idx, d in enumerate(deviations, 1):
                ws.append([
                    idx,
                    d.get("category", ""),
                    d.get("severity", ""),
                    d.get("source", ""),
                    d.get("title", ""),
                    d.get("description", ""),
                    d.get("location", ""),
                    d.get("reference_standard", ""),
                    d.get("reference_value", ""),
                    d.get("actual_value", ""),
                    d.get("action_required", ""),
                    d.get("action_taken", ""),
                    d.get("verification_status", ""),
                    d.get("notes", "")
                ])

            sheets_created += 1

    if sheets_created == 0:
        raise HTTPException(status_code=400, detail="No data to export")

    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"PID_Analysis_{export_type}_{drawing_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Type": export_type,
            "X-Session-Id": session_id
        }
    )


# =====================
# Summary
# =====================

@router.get("/{session_id}/summary")
async def get_pid_features_summary(session_id: str):
    """
    P&ID 분석 기능 요약
    """
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    def count_by_status(items: List[Dict]) -> Dict[str, int]:
        counts = {"pending": 0, "approved": 0, "rejected": 0, "modified": 0, "total": len(items)}
        for item in items:
            status = item.get("verification_status", "pending")
            if status in counts:
                counts[status] += 1
        return counts

    valves = session.get("pid_valves", [])
    equipment = session.get("pid_equipment", [])
    checklist_items = session.get("pid_checklist_items", [])
    deviations = session.get("pid_deviations", [])

    return {
        "session_id": session_id,
        "features": session.get("features", []),
        "valve_detection": {
            "detected": len(valves) > 0,
            "counts": count_by_status(valves)
        },
        "equipment_detection": {
            "detected": len(equipment) > 0,
            "counts": count_by_status(equipment)
        },
        "design_checklist": {
            "detected": len(checklist_items) > 0,
            "counts": count_by_status(checklist_items)
        },
        "deviation_analysis": {
            "detected": len(deviations) > 0,
            "counts": count_by_status(deviations),
            "by_severity": {
                sev: sum(1 for d in deviations if d.get("severity") == sev)
                for sev in ["critical", "high", "medium", "low", "info"]
            }
        },
        "overall_progress": {
            "total_items": len(valves) + len(equipment) + len(checklist_items) + len(deviations),
            "verified_items": sum(
                1 for items in [valves, equipment, checklist_items, deviations]
                for item in items
                if item.get("verification_status") != "pending"
            )
        }
    }

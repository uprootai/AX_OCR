"""P&ID Export Router

Excel 내보내기 엔드포인트
"""

import logging
from io import BytesIO
from datetime import datetime
from typing import List, Dict

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)

router = APIRouter(tags=["P&ID Export"])

# 서비스 의존성
_session_service = None


def set_session_service(session_service):
    """서비스 주입"""
    global _session_service
    _session_service = session_service


def get_session_service():
    if _session_service is None:
        raise HTTPException(status_code=500, detail="Session service not initialized")
    return _session_service


@router.post("/{session_id}/export")
async def export_to_excel(
    session_id: str,
    export_type: str = Query(..., description="내보내기 타입 (valve, equipment, checklist, all)"),
    project_name: str = Query(default="Unknown Project", description="프로젝트명"),
    drawing_no: str = Query(default="N/A", description="도면 번호"),
    include_rejected: bool = Query(default=False, description="거부된 항목 포함")
):
    """
    Excel 내보내기
    """
    session_service = get_session_service()
    session = session_service.get_session(session_id)

    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def apply_header_style(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    sheets_created = 0

    # Valve 시트
    if export_type in ["valve", "all"]:
        valves = session.get("pid_valves", [])
        if not include_rejected:
            valves = [v for v in valves if v.get("verification_status") != "rejected"]

        if valves:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Valve List"

            ws.append(["No", "Valve ID", "Type", "Category", "Region", "Confidence", "Status", "Notes"])
            apply_header_style(ws)

            for idx, v in enumerate(valves, 1):
                ws.append([
                    idx,
                    v.get("valve_id", ""),
                    v.get("valve_type", ""),
                    v.get("category", ""),
                    v.get("region_name", ""),
                    round(v.get("confidence", 0), 3),
                    v.get("verification_status", ""),
                    v.get("notes", "")
                ])

            sheets_created += 1

    # Equipment 시트
    if export_type in ["equipment", "all"]:
        equipment = session.get("pid_equipment", [])
        if not include_rejected:
            equipment = [e for e in equipment if e.get("verification_status") != "rejected"]

        if equipment:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Equipment List"

            ws.append(["No", "Tag", "Type", "Description", "Vendor Supply", "Confidence", "Status", "Notes"])
            apply_header_style(ws)

            for idx, e in enumerate(equipment, 1):
                ws.append([
                    idx,
                    e.get("tag", ""),
                    e.get("equipment_type", ""),
                    e.get("description", ""),
                    "Yes" if e.get("vendor_supply") else "No",
                    round(e.get("confidence", 0), 3),
                    e.get("verification_status", ""),
                    e.get("notes", "")
                ])

            sheets_created += 1

    # Checklist 시트
    if export_type in ["checklist", "all"]:
        checklist_items = session.get("pid_checklist_items", [])
        if not include_rejected:
            checklist_items = [c for c in checklist_items if c.get("verification_status") != "rejected"]

        if checklist_items:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Design Checklist"

            ws.append(["No", "Category", "Description", "Auto Status", "Final Status", "Evidence", "Reviewer Notes"])
            apply_header_style(ws)

            for c in checklist_items:
                ws.append([
                    c.get("item_no", ""),
                    c.get("category", ""),
                    c.get("description", ""),
                    c.get("auto_status", ""),
                    c.get("final_status", ""),
                    c.get("evidence", ""),
                    c.get("reviewer_notes", "")
                ])

            sheets_created += 1

    # Deviation 시트
    if export_type in ["deviation", "all"]:
        deviations = session.get("pid_deviations", [])
        if not include_rejected:
            deviations = [d for d in deviations if d.get("verification_status") != "rejected"]

        if deviations:
            ws = wb.active if sheets_created == 0 else wb.create_sheet()
            ws.title = "Deviation List"

            ws.append([
                "No", "Category", "Severity", "Source", "Title", "Description",
                "Location", "Reference Standard", "Reference Value", "Actual Value",
                "Action Required", "Action Taken", "Status", "Notes"
            ])
            apply_header_style(ws)

            for idx, d in enumerate(deviations, 1):
                ws.append([
                    idx,
                    d.get("category", ""),
                    d.get("severity", ""),
                    d.get("source", ""),
                    d.get("title", ""),
                    d.get("description", ""),
                    d.get("location", ""),
                    d.get("reference_standard", ""),
                    d.get("reference_value", ""),
                    d.get("actual_value", ""),
                    d.get("action_required", ""),
                    d.get("action_taken", ""),
                    d.get("verification_status", ""),
                    d.get("notes", "")
                ])

            sheets_created += 1

    if sheets_created == 0:
        raise HTTPException(status_code=400, detail="No data to export")

    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"PID_Analysis_{export_type}_{drawing_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Type": export_type,
            "X-Session-Id": session_id
        }
    )
